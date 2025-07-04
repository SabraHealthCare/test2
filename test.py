import pandas as pd  
pd.set_option('future.no_silent_downcasting', True)  
import numpy as np  
from datetime import datetime, timedelta,date
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows 
from openpyxl.utils import get_column_letter
import streamlit as st             
from io import BytesIO
from io import StringIO
import time
import  streamlit_tree_select
import copy
import streamlit.components.v1 as components
from calendar import monthrange,month_abbr
from authenticate import Authenticate 
import yaml 
from st_aggrid import AgGrid, GridUpdateMode
from st_aggrid.grid_options_builder import GridOptionsBuilder
from msal import ConfidentialClientApplication
import requests
from itertools import product
from pandas.errors import EmptyDataError
import pytz 
import chardet
from pandas.errors import EmptyDataError
import re
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import streamlit as st
from shareplum import Site
from shareplum import Office365
from shareplum.site import Version
from office365.runtime.auth.authentication_context import AuthenticationContext
from office365.sharepoint.client_context import ClientContext
import os
import random

#---------------------------define parameters--------------------------
st.set_page_config(
   initial_sidebar_state="expanded",  layout="wide")
st.title("Sabra HealthCare Monthly Reporting App")

sheet_name_discrepancy="Discrepancy_Review"
account_mapping_filename="Account_Mapping.xlsx"
BPC_pull_filename="BPC_Pull.csv"
entity_mapping_filename ="Entity_Mapping.csv"
discrepancy_filename="Total_Diecrepancy_Review.csv"
monthly_reporting_filename="Total monthly reporting.csv"
operator_list_filename="Operator_list.csv"
BPC_account_filename="Sabra_account_list.csv"
previous_monthes_comparison=0
availble_unit_accounts=["A_ACH","A_IL","A_ALZ","A_SNF","A_ALF","A_BH","A_IRF","A_LTACH","A_SP_HOSP"]
month_dic_word={10:["october","oct"],11:["november","nov"],12:["december","dec"],1:["january","jan"],\
                   2:["february","feb"],3:["march","mar"],4:["april","apr"],\
                   5:["may"],6:["june","jun"],7:["july","jul"],8:["august","aug"],9:["september","sep"]}
month_dic_num={10:["10/","-10","/10","10","_10"],11:["11/","-11","/11","11","_11"],12:["12/","-12","/12","12"],1:["01/","1/","-1","-01","/1","/01"],\
                   2:["02/","2/","-2","-02","/2","/02"],3:["03/","3/","-3","-03","/3","/03"],4:["04/","4/","-4","-04","/4","/04"],\
                   5:["05/","5/","-5","-05","/5","/05"],6:["06/","6/","-06","-6","/6","/06"],\
                   7:["07/","7/","-7","-07","/7","/07"],8:["08/","8/","-8","-08","/8","/08"],9:["09/","9/","-09","-9","/9","/09"]}
year_dic={2024:["2024","24"],2025:["2025","25"],2026:["2026","26"]} 	
month_map = {"Jan": "01", "Feb": "02", "Mar": "03", "Apr": "04", "May": "05", "Jun": "06", "Jul": "07", "Aug": "08","Sep": "09", "Oct": "10", "Nov": "11", "Dec": "12"}
PL_total_names=["Total Patient Days in P&L","Total Revenue in P&L","Total OPEX in P&L","Total Expense in P&L"]
#One drive authority. Set application details
client_id = 'bc5f9d8d-eb35-48c3-be6d-98812daab3e3'
client_secret='PgR8Q~HZE2q-dmOb2w_9_0VuxfT9VMLt_Lp3Jbce'
tenant_id = '71ffff7c-7e53-4daa-a503-f7b94631bd53'
authority = 'https://login.microsoftonline.com/' + tenant_id
user_id= '62d4a23f-e25f-4da2-9b52-7688740d9d48'  # shali's user id of onedrive
PL_path="Documents/Tenant Monthly Uploading/Tenant P&L"
mapping_path="Documents/Tenant Monthly Uploading/Tenant Mapping"
master_template_path="Documents/Tenant Monthly Uploading/Master Template"

# SharePoint credentials and site details
SHAREPOINT_URL = "https://sabrahealthcare.sharepoint.com"  # Full URL with scheme
SHAREPOINT_SITE = "https://sabrahealthcare.sharepoint.com/sites/S-Cloud"  # Full site URL
sharepoint_username = "sli@sabrahealth.com"  # Replace with your SharePoint username
sharepoint_password = "June2023SL!"
email_body=""
if "email_body_for_Sabra" not in st.session_state:
    st.session_state.email_body_for_Sabra = ""  # only email these info to sabra

today= datetime.now(pytz.timezone('America/Los_Angeles')).date()
current_year= today.year
current_month= today.month
current_day=today.day  
operators_remove_hidden_rowcol=["Ignite","Recovery Centers of America","Nexus Systems"]#"Creative Solutions"

# Acquire a token using client credentials flow
app = ConfidentialClientApplication(
    client_id,
    authority=authority,
    client_credential=client_secret)

token_response = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
access_token = token_response['access_token']
headers = {'Authorization': 'Bearer ' + access_token,}    

account_mapping_str_col=["Tenant_Account",]
entity_mapping_str_col=["DATE_ACQUIRED","DATE_SOLD_PAYOFF","Sheet_Name_Finance","Sheet_Name_Occupancy","Sheet_Name_Balance_Sheet","Column_Name"]

def sanitize_filename(filename):
    """Sanitize only the base name of the file, keeping extension like .pdf"""
    name, ext = os.path.splitext(filename)

    # Remove apostrophes, replace &, remove other illegal SharePoint characters
    name = name.replace("'", "")
    name = name.replace("&", "and")
    name = re.sub(r"[#%*:<>\?/\\{|}\"\.]", "", name)  # remove bad chars from name (keep dot only in ext)

    return f"{name}{ext}"

def Upload_To_Sharepoint(files, sharepoint_folder, new_file_names):
    try:
        # Authenticate with SharePoint
        authcookie = Office365(SHAREPOINT_URL, username=sharepoint_username, password=sharepoint_password).GetCookies()
        site = Site(SHAREPOINT_SITE, version=Version.v365, authcookie=authcookie)

        sharepoint_folder_obj = site.Folder(sharepoint_folder)
        success_files = []
        failed_files = []  

        for i, file in enumerate(files):
            original_name = new_file_names[i]
            safe_name = sanitize_filename(original_name)  # Sanitize for SharePoint
            temp_file_path = os.path.join(".", safe_name)

            try:   
                # Save uploaded file temporarily
                with open(temp_file_path, "wb") as f:
                    f.write(file.getbuffer())

                # Upload to SharePoint
                with open(temp_file_path, "rb") as file_content:
                    sharepoint_folder_obj.upload_file(file_content, safe_name)

                success_files.append(file.name)

            except:
                failed_files.append(file.name)

            finally:
                if os.path.exists(temp_file_path):
                    os.remove(temp_file_path)

        if len(success_files) == len(files):
            return True, success_files
        else:
            return False, failed_files

    except:
        return False, []

 
def Send_Confirmation_Email(receiver_email_list, subject, email_body):
    username = 'sabrahealth.com'  
    password = 'b1bpwmzxs9hnbpkM'  #SMTP2GO password, not the API_key

    # Create the email
    msg = MIMEMultipart('mixed')
    msg['Subject'] = subject
    msg['From'] = "Sabra_reporting@sabrahealth.com"
    msg['To'] =  ", ".join(receiver_email_list)
    
    html_part = MIMEText(email_body, 'html')
    # Attach both plain text and HTML messages
    msg.attach(html_part)

    # Connect to SMTP2GO server and send email
    try:
        mailServer = smtplib.SMTP('mail.smtp2go.com', 2525)  # Can also use 8025, 587, or 25
        mailServer.ehlo()
        mailServer.starttls()
        mailServer.ehlo()
        mailServer.login(username, password)
        mailServer.sendmail("Sabra_reporting@sabrahealth.com", receiver_email_list, msg.as_string())
        mailServer.close()
    except Exception as e:
        st.write( f"Failed to send confirmation email.")
	    
#directly save the uploaded (.xlsx) file to onedrive
def Upload_To_Onedrive(uploaded_files,path,file_names):
    i=0
    success_files = []
    failed_files = [] 
    try:
        for uploaded_file in uploaded_files:
            file_name=file_names[i]
            i+=1
            # Set the API endpoint and headers
            api_url = f'https://graph.microsoft.com/v1.0/users/{user_id}/drive/items/root:/{path}/{file_name}:/content'

            # Ensure the file pointer is at the start
            uploaded_file.seek(0)
    
            # Read the file content as binary data
            file_content = uploaded_file.read()
    
            # Set the Content-Type header for Excel files
            headers.update({'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'})
    
            # Make the request to upload the file
            response = requests.put(api_url, headers=headers, data=file_content)
	
            if response.status_code in [200,201]:# or response.status_code==201:
                success_files.append(file_name)
                continue
            else:
                failed_files.append(file_name)
                continue
        if len(success_files) == len(uploaded_files):
            return True,success_files
        else:
            return False, failed_files
    except:
        return False,[]
# no cache read csv/excel from onedrive , return dataframe
def detect_encoding(file_content):
    result = chardet.detect(file_content)
    return result['encoding']

def Read_File_From_Onedrive(path, file_name, file_type, str_col_list=None):
    if str_col_list is None:
        str_col_list = []
    
    # Set the API endpoint and headers for file download
    api_url = f'https://graph.microsoft.com/v1.0/users/{user_id}/drive/root:/{path}/{file_name}:/content'
    
    # Make the request to download the file
    response = requests.get(api_url, headers=headers)
    
    # Check the status code 
    if response.status_code == 200 or response.status_code == 201:
        file_content = response.content
        
        try:
            # Set the dtype dictionary for specified columns
            dtype_dict = {col: str for col in str_col_list}
            
            if file_type.upper() == "CSV":    
                detected_encoding = detect_encoding(file_content)
                if file_name.lower().endswith(".csv"):
                    df = pd.read_csv(BytesIO(file_content), encoding=detected_encoding, on_bad_lines='skip', dtype=dtype_dict)
                elif file_name.lower().endswith(".xlsx"):
                    df = pd.read_excel(BytesIO(file_content), dtype=dtype_dict, engine='openpyxl')
                return df
            elif file_type.upper() == "XLSX":
                df = pd.read_excel(BytesIO(file_content), dtype=dtype_dict, engine='openpyxl')
                return df
            elif file_type.upper() == "YAML":
                config = yaml.safe_load(file_content)
                return config
            elif file_type.upper() == "VIDEO": 
                return BytesIO(response.content)

        except pd.errors.EmptyDataError:
            st.write("EmptyDataError: The file is empty.")
            return False
        except pd.errors.ParserError as e:
            st.write(f"ParserError: {e}")
            return False
        except Exception as e:
            st.write(f"Unexpected error: {e}")
            return False
        
    else: 
        st.write(f"Failed to download file: {response.status_code}")
        return False

# no cache, save a dataframe to OneDrive 
def Save_File_To_Onedrive(df, path, file_name, file_type):
    try:
        # Filter out unwanted columns
        df = df[list(filter(lambda x: x != "index" and "Unnamed:" not in x, df.columns))]
        
        # Define your Microsoft Graph API endpoint, user ID, and headers
        api_url = f'https://graph.microsoft.com/v1.0/users/{user_id}/drive/items/root:/{path}/{file_name}:/content'
        # Handle file_type
        if file_type.upper() == "CSV":
            file_name = f"{file_name}.csv" if not file_name.endswith(".csv") else file_name
            file_content = df.to_csv(index=False).encode()
        elif file_type.upper() == "XLSX":
            file_name = f"{file_name}.xlsx" if not file_name.endswith(".xlsx") else file_name
            excel_buffer = BytesIO()
            df.to_excel(excel_buffer, index=False, engine='xlsxwriter')
            excel_buffer.seek(0)
            file_content = excel_buffer.read()
        else:
            raise ValueError("Unsupported file type. Use 'CSV' or 'XLSX'.")
        
        # Send the request to OneDrive
        response = requests.put(api_url, headers=headers, data=BytesIO(file_content))
        
        # Check the response
        if response.status_code == 200:
            return True
        else:
            return False
    except Exception as e:
        st.error(f"Error: {e}")
        return False

                



# For updating account_mapping, entity_mapping, reporting_month_data, only for operator use
# if entity_list is provided,
def Update_File_Onedrive(path,file_name,new_data,operator,file_type="CSV",entity_list=None,str_col_list=None):  # replace original data
    entity_list = entity_list or []   
    original_data=Read_File_From_Onedrive(path,file_name,file_type,str_col_list)
    new_data=new_data.reset_index(drop=False)
	
    if  isinstance(original_data, pd.DataFrame):
        if "TIME" in original_data.columns and "TIME" in new_data.columns:
            original_data.TIME = original_data.TIME.astype(str)
            months_of_new_data=new_data["TIME"].unique()
            condition = (original_data['Operator'] == operator) & (original_data['TIME'].isin(months_of_new_data))
            if entity_list:
                condition &= original_data['ENTITY'].isin(entity_list)
                new_data = new_data[new_data["ENTITY"].isin(entity_list)]
            # remove original data by operator and month
            original_data = original_data[~condition]


        else:
            condition = (original_data['Operator'] == operator)
            if entity_list:
                condition &= original_data['ENTITY'].isin(entity_list)
            original_data = original_data[~condition]

        updated_data = pd.concat([original_data, new_data])
        updated_data = updated_data.drop(columns='index', errors='ignore')
    else:
        updated_data = new_data.drop(columns='index', errors='ignore')
    return Save_File_To_Onedrive(updated_data,path,file_name,file_type)  # return True False

def Format_Value(column):
    def format_value(x):
        if pd.isna(x) or (isinstance(x, str) and x.strip() == "") or x == 0:
            return None
        elif isinstance(x, float):
            return round(x, 1)
        return x
    return column.apply(format_value)

# Function to update the value in session state
def clicked(button_name):
    st.session_state.clicked[button_name] = True

# No cache
def Initial_Mapping(operator):
    BPC_pull=Read_File_From_Onedrive(mapping_path,BPC_pull_filename,"CSV")
    BPC_pull = (BPC_pull[BPC_pull["Operator"] == operator]
            .set_index(["ENTITY", "Sabra_Account"])
            .dropna(axis=1, how='all')
            .rename(columns=str))
    #st.write("BPC_pull",BPC_pull)
    # Read account mapping file from OneDrive
    account_mapping_all = Read_File_From_Onedrive(mapping_path,account_mapping_filename,"XLSX",account_mapping_str_col)
    #st.write("account_mapping_filename",account_mapping_filename,"mapping_path",mapping_path,"account_mapping_all",account_mapping_all)    
    # Handle case where there's only one row and it corresponds to a template
    account_mapping = account_mapping_all[account_mapping_all["Operator"]==operator].copy()
    #st.write(account_mapping)
    #st.write(account_mapping["Sabra_Account"].values[0])
    if account_mapping.shape[0] == 1 and account_mapping["Sabra_Account"].values[0] == 'TEMPLATE':
        account_mapping = account_mapping_all[account_mapping_all["Operator"] == "Template"].copy()
        account_mapping["Operator"] = operator	
    
    #st.write("account_mapping1",account_mapping)  
    # Clean and format account mapping columns
    account_mapping_cols = ["Sabra_Account", "Sabra_Second_Account"]
    account_mapping.loc[:, account_mapping_cols] = account_mapping[account_mapping_cols].apply(
    lambda col: col.map(lambda x: x.upper().strip() if pd.notna(x) else x)
)

    account_mapping.loc[:, "Tenant_Account"] = account_mapping["Tenant_Account"].map(
    lambda x: str(int(x)).strip().upper()
    if pd.notna(x) and isinstance(x, float)
    else (str(x).strip().upper() if pd.notna(x) else x)
)

    if "Category" in account_mapping.columns:
        account_mapping = account_mapping.drop(columns="Category")
    account_mapping=account_mapping.merge(BPC_Account[["BPC_Account_Name","Category"]], left_on="Sabra_Account",right_on="BPC_Account_Name",how="left").drop(columns="BPC_Account_Name")
    account_mapping = account_mapping[["Operator", "Sabra_Account", "Sabra_Second_Account", "Tenant_Account", "Conversion","Category"]]
    #st.write("account_mapping",account_mapping)  

    entity_mapping = (Read_File_From_Onedrive(mapping_path, entity_mapping_filename, "CSV", entity_mapping_str_col)
                  .reset_index(drop=True)
                  .query("Operator == @operator") 
                  .set_index("ENTITY"))
 
    entity_mapping[["DATE_ACQUIRED", "DATE_SOLD_PAYOFF"]] = entity_mapping[["DATE_ACQUIRED", "DATE_SOLD_PAYOFF"]].astype(str)  
    #st.write("entity_mapping",entity_mapping)
    for col in ["Sheet_Name_Finance","Sheet_Name_Occupancy","Sheet_Name_Balance_Sheet","Column_Name"]:
        entity_mapping.loc[:, col] = entity_mapping[col].apply(lambda x: x.replace("'","") if pd.notna(x) else x)

    #st.write("entity_mapping",entity_mapping)
    return BPC_pull,entity_mapping,account_mapping

	
# Intialize a list of tuples containing the CSS styles for table headers
th_props = [('font-size', '14px'), ('text-align', 'left'),
            ('font-weight', 'bold'),('color', '#6d6d6d'),
            ('background-color', '#eeeeef'), ('border','1px solid #eeeeef')]

# Intialize a list of tuples containing the CSS styles for table data
td_props = [('font-size', '14px'), ('text-align', 'left')]

# Aggregate styles in a list
styles = [dict(selector="th", props=th_props),dict(selector="td", props=td_props)]


def left_align(s, props='text-align: left;'):
    return props
	
css='''
<style>
    section.main>div {
        padding-bottom: 1rem;
    }
    [data-testid="table"]>div>div>div>div>div {
        overflow: auto;
        height: 20vh;
    }
</style>
'''
st.markdown(css, unsafe_allow_html=True)

@st.cache_data
def Create_Tree_Hierarchy():
    # Initialize hierarchy with default options
    parent_hierarchy_main = [{'label': "No need to map", 'value': "No need to map"}]
    parent_hierarchy_second = []
    
    # Read account data
    BPC_Account = Read_File_From_Onedrive(mapping_path, BPC_account_filename, "CSV")
    
    # Function to create hierarchy for a given type
    def create_hierarchy(account_type):
        hierarchy = []
        for category in BPC_Account[BPC_Account["Type"] == account_type]["Category"].unique():
            children_hierarchy = [
                {'label': account, 
                 'value': BPC_Account.loc[(BPC_Account["Sabra_Account_Full_Name"] == account) & (BPC_Account["Type"] == account_type), "BPC_Account_Name"].item()}
                for account in BPC_Account[(BPC_Account["Category"] == category) & (BPC_Account["Type"] == account_type)]["Sabra_Account_Full_Name"]
            ]
            hierarchy.append({'label': category, 'value': category, 'children': children_hierarchy})
        return hierarchy
    
    # Create hierarchies for main and second types
    parent_hierarchy_main += create_hierarchy("Main")
    parent_hierarchy_second += create_hierarchy("Second")
    
    # Select relevant columns for the output
    BPC_Account = BPC_Account[["BPC_Account_Name", "Sabra_Account_Full_Name", "Category"]]
    return parent_hierarchy_main, parent_hierarchy_second, BPC_Account

parent_hierarchy_main,parent_hierarchy_second,BPC_Account=Create_Tree_Hierarchy()

#-----------------------------------------------functions---------------------------------------------
def ChangeWidgetFontSize(wgt_txt, wch_font_size = '12px'):
    htmlstr = """<script>var elements = window.parent.document.querySelectorAll('*'), i;
                    for (i = 0; i < elements.length; ++i) { if (elements[i].innerText == |wgt_txt|) 
                        { elements[i].style.fontSize='""" + wch_font_size + """';} } </script>  """
    htmlstr = htmlstr.replace('|wgt_txt|', "'" + wgt_txt + "'")
    components.html(f"{htmlstr}", height=0, width=0)


# Parse the df and get filter widgets based for provided columns
		
def Identify_Tenant_Account_Col(PL, sheet_name, sheet_type_name, account_pool, pre_max_match_col):
    #st.write("PL",PL,"account_pool",account_pool)

    def get_match_count(col_index):
        candidate_col = PL.iloc[:, col_index].apply(lambda x: str(int(x)).strip().upper() \
				if pd.notna(x) and isinstance(x, float) else (str(x).strip().upper() if pd.notna(x) else x))
        #st.write("candidate_col",candidate_col)
        non_empty_col = candidate_col[candidate_col != '']
        match_count = sum(candidate_col.isin(account_pool))
        #st.write("match_count",match_count)
        return match_count, len(non_empty_col)
    
    # Check the pre-identified columns first
    if pre_max_match_col != [10000] and pre_max_match_col[0] < PL.shape[1] and len(pre_max_match_col)==1:
        for i in range(len(pre_max_match_col)):
            match_count, non_empty_count = get_match_count(pre_max_match_col[i])
            if match_count > 0 and (match_count > 1 or match_count / non_empty_count > 0.2):
                if i == len(pre_max_match_col)-1:
                    #st.write("_______________________________use pre_max_match_col_______________________:",pre_max_match_col)
                    return pre_max_match_col
    
    # If pre-identified columns are not sufficient, search for potential matches across the first 15 columns
    match_counts = []
    for col in range(min(15, PL.shape[1])):
        match_count, _ = get_match_count(col)
        match_counts.append((match_count, col))
    #st.write("match_counts",match_counts)
    # Sort by match count in descending order
    match_counts.sort(reverse=True, key=lambda x: x[0])
    
    # Return the top columns with the highest match counts
    top_matches = [match[1] for match in match_counts if match[0] > 0]
    if len(top_matches)>0:
        return top_matches # return a list of col index
    
    # If no match is found, raise an error
    st.error(f"Failed to identify tenant account columns in {sheet_type_name} sheet —— {sheet_name}")
    st.stop()

def download_report(df,button_display):
    download_file=df.to_csv(index=False).encode('utf-8')
    return st.download_button(label="Download "+button_display,data=download_file,file_name=button_display+".csv",mime="text/csv")
 
def Get_Year(single_string):
    for Year, keywords in year_dic.items():
        for keyword  in keywords: # keywards are "2024","24"
            if re.search(re.escape(keyword), single_string):
                return Year,keyword
    return 0,0

def Get_Month_Year(single_string):
    if pd.isna(single_string):
        return 0,0
    if isinstance(single_string, datetime):
        return int(single_string.month),int(single_string.year)
    if isinstance(single_string, (int,float)) and single_string not in year_dic:
        return 0,0
    
    single_string=str(single_string).lower()
    year,year_num=Get_Year(single_string)

    if year!=0:
        single_string=single_string.replace(year_num,"")
        if not single_string:
            return 0,year
    single_string=single_string.replace("30","").replace("31","").replace("29","").replace("28","")
    #st.write("single_string1",single_string)
    for month_i ,month_words in month_dic_word.items():#[10,11,3...12]
        for  month_word in month_words: # month_word is element of ['december','dec',"nov",...]
            if month_word in single_string:  # month is words ,like Jan Feb... year is optional
                remaining=single_string.replace(month_word,"").replace("/","").replace("-","").replace(" ","").replace("_","")
                remaining=remaining.replace("asof","").replace("actual","").replace("mtd","")
                #st.write("remaining",remaining)
                #if there are more than 3 other char left in the string, this string is not month 
                if len(remaining)<3:  
                    return month_i,year
			
    # didn't detect month words in above code, check number format: 3/31/2024, 3/2023...
    # if there is no year, skip
    if year==0:
        return 0,0   
        
    for month_i, month_nums  in month_dic_num.items(): 
        for month_num in month_nums:
            if month_num in single_string:  # month is number ,like 01/, 02/,   year is Mandatory
                remaining=single_string.replace(month_num,"").replace("/","").replace("-","").replace(" ","").replace("_","")
                remaining=remaining.replace("asof","").replace("actual","").replace("mtd","")
                #if there are more than 3 other char in the string, this string is not month 
                if len(remaining)<3: 
                    return month_i,year	
    # didn't find month. return month as 0
    return 0,0   

# add year to month_header: identify current year/last year giving a list of month
def Fill_Year_To_Header(PL,month_row_index,full_month_header,sheet_name,reporting_month):
    # month_row_index is the row number for month header

    #remove rows with nan tenant account
    nan_index = list(filter(lambda x: pd.isna(x) or str(x).strip().lower() in ["nan", "", "0"], PL.index))
    PL_filtered = PL.drop(nan_index)
    column_mask = [all(val == 0 or not isinstance(val, (int, float)) or pd.isna(val) for val in PL_filtered.iloc[:, i]) for i in range(PL_filtered.shape[1])]
    
  # Apply the mask to set these columns to NaN in the row specified by month_row_index
    full_month_header = [0 if column_mask[i] else full_month_header[i] for i in range(len(full_month_header))]

    month_list=list(filter(lambda x:x!=0,full_month_header))
    month_len=len(month_list)
    # Initialize the full_year_header with zeros
    full_year_header=[0] * len(full_month_header)
    if month_len==1:
        year=reporting_month[0:4]
        PL_date_header= [f"{year}{month:02d}" if month!=0 else 0 for month in full_month_header]
        return PL_date_header
	    
    add_year=month_list
    report_year=int(reporting_month[:4])
    last_report_year=report_year-1
    year_change=0  

    # Check for ascending or descending order in the month list
    inv=[int(month_list[month_i+1])-int(month_list[month_i]) for month_i in range(month_len-1) ]
    ascending_check=sum([x in [1,-11] for x in inv])
    descending_check=sum([x in [-1,11] for x in inv])

    # Convert reporting month to date for comparison	
    reporting_month_date=datetime.strptime(str(reporting_month[4:6])+"/01/"+str(report_year),'%m/%d/%Y').date()   
    # Handle descending months, month_list[0]<today.month 
    if descending_check>0 and descending_check>ascending_check: 
        date_of_assumption=datetime.strptime(str(month_list[0])+"/01/"+str(report_year),'%m/%d/%Y').date()
        if date_of_assumption==reporting_month_date:	
            report_year_start=report_year
        elif date_of_assumption<today and date_of_assumption.month<today.month:
            report_year_start=report_year
        elif date_of_assumption>=today:
            report_year_start=last_report_year
        for i in range(month_len):
            add_year[i]=report_year_start-year_change
            if i<month_len-1 and add_year[i+1]==12:
                year_change+=1
            
    # month ascending  
    elif ascending_check>0 and ascending_check> descending_check: 
        date_of_assumption=datetime.strptime(str(month_list[-1])+"/01/"+str(report_year),'%m/%d/%Y').date() 
        if date_of_assumption==reporting_month_date:
            report_year_start=report_year
        elif date_of_assumption<today:
            report_year_start=report_year
        elif date_of_assumption>=today:
            report_year_start=last_report_year
        for i in range(-1,month_len*(-1)-1,-1):
            add_year[i]=report_year_start-year_change
            if i>month_len*(-1) and add_year[i-1]==12:
                year_change+=1
    #    # Handle other cases and errors  , month decending 	    
    elif (month_list[0]>month_list[1] and month_list[0]!=12) or (month_list[0]==1 and month_list[1]==12):
        date_of_assumption=datetime.strptime(str(month_list[0])+"/01/"+str(report_year),'%m/%d/%Y').date()
        if date_of_assumption<today and date_of_assumption.month<today.month:
            report_year_start=report_year
        elif date_of_assumption>=today:
            report_year_start=int(report_year)-1
        for i in range(month_len):
            add_year[i]=report_year_start-year_change
            if i<month_len-1 and add_year[i+1]==12:
                year_change+=1
     # month ascending
    elif (month_list[0]<month_list[1] and month_list[0]!=12) or (month_list[0]==12 and month_list[1]==1): 
        date_of_assumption=datetime.strptime(str(month_list[-1])+"/01/"+str(report_year),'%m/%d/%Y').date()    
        if date_of_assumption<today:
            report_year_start=report_year
        elif date_of_assumption>=today:
            report_year_start=int(report_year)-1
        for i in range(-1,month_len*(-1)-1,-1):
            add_year[i]=report_year_start-year_change
            if i>month_len*(-1) and add_year[i-1]==12:
                year_change+=1
    else:
        st.error("Fail to identify Year in sheet {}, please add the year for the month and re-upload.".format(sheet_name))
        st.stop()
    j=0
  
    for i in range(len(full_month_header)):
        if full_month_header[i]!=0:
            full_year_header[i]=add_year[j]
            j+=1
    PL_date_header= [f"{year}{month:02d}" if year!=0 else 0 for year, month in zip(full_year_header, full_month_header)]
    return PL_date_header
	
@st.cache_data 
def Check_Available_Units(reporting_month_data,Total_PL,check_patient_days,reporting_month,email_body):
    #check patient days,fill missing operating beds to reporting_month_data
    month_days=monthrange(int(reporting_month[:4]), int(reporting_month[4:]))[1]
    problem_properties=[]
    properties_fill_Aunit=[]
    zero_patient_days=[]
    total_property_list=reporting_month_data["Property_Name"].unique()
    error_for_email=""

    for property_i in total_property_list:
        # Initialize values to 0
        patient_day_i = 0
        operating_beds_i = 0

        # Try to retrieve patient days if it exists in the index
        if (property_i, "Patient Days") in check_patient_days.index:
            patient_day_i = check_patient_days.loc[(property_i, "Patient Days"), reporting_month]

        # Try to retrieve operating beds if it exists in the index
        if (property_i, "Operating Beds") in check_patient_days.index:
            operating_beds_i = check_patient_days.loc[(property_i, "Operating Beds"), reporting_month]


        # Perform validations
        if patient_day_i > 0 and operating_beds_i * month_days > patient_day_i:
            continue
        elif operating_beds_i > 0 and patient_day_i > operating_beds_i * month_days:
            error_message = (\
            f"The number of patient days for {property_i} exceeds its available days "\
            f"(Operating Beds * {month_days}). This will result in incorrect occupancy." )
            st.error("Error: " + error_message)
            problem_properties.append(property_i)
            error_for_email += f"<li>{error_message}</li>"
      
        elif operating_beds_i==0 and patient_day_i==0:
            zero_patient_days.append(property_i)
        elif patient_day_i==0 and operating_beds_i>0:
            error_message="{} is missing patient days. If this facility is not currently functioning or in operation, please remove the number of operating beds associated with it.".format(property_i)
            st.error("Error: "+error_message)
            problem_properties.append(property_i)   
            error_for_email+="<li> "+error_message+"</li>"
        elif patient_day_i>0 and operating_beds_i==0:
            properties_fill_Aunit.append(property_i)
		
    if len(properties_fill_Aunit)>0 and not BPC_pull.empty:    
        BPC_pull_reset = BPC_pull.reset_index()
        previous_A_unit = BPC_pull_reset.loc[(BPC_pull_reset["Sabra_Account"].str.startswith("A_")) &(BPC_pull_reset["Property_Name"].isin(properties_fill_Aunit)),["ENTITY","Property_Name","Sabra_Account","A_unit"]]
        previous_A_unit=previous_A_unit.merge(BPC_Account, left_on="Sabra_Account", right_on="BPC_Account_Name",how="left")	
        previous_A_unit=previous_A_unit.rename(columns={"A_unit":reporting_month})
        reporting_month_data  = pd.concat([reporting_month_data, previous_A_unit], axis=0)
	# delete oiginal A_ account in reporting_month_data if original A_ account value is 0 or None
        reporting_month_data = reporting_month_data[\
                            ~((reporting_month_data[reporting_month].isin([0, None])) & reporting_month_data["BPC_Account_Name"].str.startswith("A_"))]

        Total_PL=pd.concat([Total_PL, previous_A_unit.set_index(["ENTITY","Sabra_Account"])[reporting_month]], axis=0)
        # delete the original operating beds if they are 0 or none. otherwise there will be two A_SNF, one has value 0
        Total_PL = Total_PL[~((Total_PL[reporting_month].isin([0, None])) \
			      & (Total_PL.index.get_level_values("Sabra_Account").str.startswith("A_")))]
	    
        if previous_A_unit.shape[0]>1:
            st.warning("The following properties are missing operating beds. Historical data has been used to fill in the gaps. If this information is incorrect, please update the operating beds in the P&L and re-upload.")
        elif previous_A_unit.shape[0]==1:
            st.error("{} is missing operating beds. Historical data has been used to fill in the missing info as shown below. If this data is incorrect, please add the operating beds and re-upload P&L.".format(properties_fill_Aunit[0]))
        previous_A_unit_display = previous_A_unit.pivot(index=["Sabra_Account"], columns="Property_Name", values=reporting_month)
        st.write(previous_A_unit_display) 
	    
        # check the filled operating beds and corresponding patient days    
        for property_i in properties_fill_Aunit:
            # Initialize patient_day_i and operating_beds_i to 0
            patient_day_i = 0
            operating_beds_i = 0
    
            # retrieve patient days
            if (property_i, "Patient Days") in check_patient_days.index:
                patient_day_i = check_patient_days.loc[(property_i, "Patient Days"), reporting_month]
    
            # retrieve operating beds
            if property_i in previous_A_unit["Property_Name"].values:
                operating_beds_i = previous_A_unit.loc[previous_A_unit["Property_Name"] == property_i, reporting_month].sum()
    
            if patient_day_i>0 and operating_beds_i*month_days>patient_day_i:
                continue
            elif patient_day_i>0 and operating_beds_i==0:
                error_message="{} is missing operating beds. Please add its operating beds and re-upload.".format(property_i)
                st.error("Error："+error_message)
                problem_properties.append(property_i)
                error_for_email+="<li> "+error_message+"</li>"
            elif operating_beds_i>0 and patient_day_i>operating_beds_i*month_days:
                error_message="The number of patient days for {} exceeds its available days (Operating Beds * {}). This will result in incorrect occupancy.".format(property_i,month_days)		
                st.error("Error："+error_message)
                problem_properties.append(property_i)
                error_for_email+="<li> "+error_message+"</li>"
                check_patient_days.loc[(property_i,"Operating Beds"),reporting_month]=operating_beds_i
	
    if len(problem_properties)>0:
        check_patient_days_display=check_patient_days.loc[(problem_properties,slice(None)),reporting_month].reset_index(drop=False)
        check_patient_days_display=check_patient_days_display.pivot_table(index=["Property_Name"],columns="Category", values=reporting_month,aggfunc='last').astype(int)  
        check_patient_days_display.reset_index(inplace=True)  
        if "Operating Beds" not in check_patient_days_display.columns:
            check_patient_days_display["Operating Beds"]=0

        check_patient_days_display.columns.name=None
        check_patient_days_display=check_patient_days_display.rename(columns={"Property_Name": "Property"})
        st.dataframe(check_patient_days_display.style.map(color_missing, subset=["Patient Days","Operating Beds"]).format(precision=0, thousands=","),hide_index=True)
        email_body+= f" <p>Please pay attention to the improper entries in the patient days:</p>{check_patient_days_display.to_html(index=False)}"+"<ul>"+error_for_email+"</ul>"	
    return reporting_month_data,Total_PL,email_body

@st.cache_data  
def Identify_Month_Row(PL,tenant_account_col_values,tenantAccount_col_no,sheet_name,sheet_type,pre_date_header): 

    #st.write("sheet_name",sheet_name)
    #pre_date_header is the date_header from last PL. in most cases all the PL has same date_header, so check it first
    #st.write("pre_date_header",pre_date_header)
    if len(pre_date_header[2])!=0:
        if PL.iloc[pre_date_header[1],:].equals(pre_date_header[2]):
            return pre_date_header
    PL_col_size=PL.shape[1]

    # Create a set of tenant accounts that need mapping
    accounts_to_map = [account for account, sabra_account in zip(account_mapping['Tenant_Account'], account_mapping['Sabra_Account']) if sabra_account!= 'NO NEED TO MAP']
    #st.write("tenant_account_col_values",tenant_account_col_values,"accounts_to_map",accounts_to_map)
    # Create a boolean mask using a list comprehension
    tenant_account_row_mask = [account in accounts_to_map for account in tenant_account_col_values]
    #st.write("tenant_account_row_mask",tenant_account_row_mask)	
    #first_tenant_account_row is the row number for the first tenant account (except for no need to map)
    #st.write("tenant_account_row_mask",tenant_account_row_mask)
    if not any(tenant_account_row_mask):  #all the accounts in tenant_account_col are new accounts 
        PL_temp=PL.copy()
        first_tenant_account_row=PL_temp.shape[0]
    else:
        PL_temp=PL.loc[tenant_account_row_mask]
        first_tenant_account_row=tenant_account_row_mask.index(max(tenant_account_row_mask))
        #st.write("tenantAccount_col_no",first_tenant_account_row)
    #valid_col_mask labels all the columns as ([False, False, True,.True..False...])
    #1. on the right of tenantAccount_col_no 
    #2.contain numeric value 
    #3. not all 0 or nan in tenant_account_row. 

    valid_col_mask = PL_temp.apply(lambda x: ( pd.to_numeric(x, errors='coerce').notna().any() and \
           not all((v == 0 or pd.isna(v) or isinstance(v, str) or not isinstance(v, (int, float))) for v in x)\
         ) if PL_temp.columns.get_loc(x.name) > tenantAccount_col_no else False, axis=0)
    valid_col_index=[i for i, mask in enumerate(valid_col_mask) if mask]
    #st.write("PL_temp",PL_temp,"valid_col_mask",valid_col_mask,valid_col_index)
    if len(valid_col_index)==0: # there is no valid data column
        st.write("Didn't detect any data in sheet {}".format(sheet_name))
        return [],0,[]
    # nan_num_column is the column whose value is nan or 0 for PL.drop(nan_index)
    #nan_num_column = [all(val == 0 or pd.isna(val) or not isinstance(val, (int, float)) for val in PL.drop(nan_index).iloc[:, i]) for i in range(PL.drop(nan_index).shape[1])]
    month_table=pd.DataFrame(0,index=range(first_tenant_account_row), columns=range(PL_col_size))
    year_table=pd.DataFrame(0,index=range(first_tenant_account_row), columns=range(PL_col_size))
    for row_i in range(first_tenant_account_row): # only search month/year above the first tenant account row
        for col_i in valid_col_index:  # only search the columns that contain numberic data and on the right of tenantAccount_col_no
            month_table.iloc[row_i,col_i],year_table.iloc[row_i,col_i]=Get_Month_Year(PL.iloc[row_i,col_i]) 
    max_len=0
    candidate_date=[]
    month_count = month_table.apply(lambda row: (row != 0).sum(), axis=1).tolist()
    #st.write("month_table",month_table)
    if not all(x==0 for x in month_count):
       # month_sort_index is the index(row number) which contain month/year, and sorted desc. month_sort_index[0] is the row number that contrain most months in PL
        non_zero_indices = [(index, month_c) for index, month_c in enumerate(month_count) if month_c!= 0]
        sorted_non_zero_indices = sorted(non_zero_indices, key=lambda x: x[1], reverse=True)
        month_sort_index = [index for index, month_c in sorted_non_zero_indices]

        for month_row_index in month_sort_index: 
            month_row=list(month_table.iloc[month_row_index,])
            month_list=list(filter(lambda x:x!=0,month_row))
            month_len=len(month_list)
            max_match_year=0
            for i in [0,1,-1]:  # identify year in corresponding month row, or above(-1) or below (+1) month row
                if month_row_index+i>=0 and month_row_index+i<year_table.shape[0]:
                    year_row=list(year_table.iloc[month_row_index+i,])
                    year_match = [year for month, year in zip(month_row, year_row) if month!= 0 and year!=0]
                  
                    if len(year_match)==month_len:
                        year_table.iloc[month_row_index,:] = [year_table.iloc[month_row_index+i,j] if month != 0 else 0 for j, month in enumerate(month_row)]
                        max_match_year=len(year_match)
                        break
                    elif len(year_match)<month_len and len(year_match)>max_match_year:
                        year_table.iloc[month_row_index,:] = [year_table.iloc[month_row_index+i,j] if month != 0 else 0 for j, month in enumerate(month_row)]
                        max_match_year=len(year_match)
                    else:
                        continue
            
            if month_count[month_row_index]>1:   # if there are more than one month in the header	    
	        #check month continuous, there are at most two types of differences in the month list which are in 1,-1,11,-11 
                inv=[int(month_list[month_i+1])-int(month_list[month_i]) for month_i in range(month_len-1) ]
                continuous_check_bool=[x in [1,-1,11,-11] for x in inv]
                len_of_continuous=sum(continuous_check_bool)
                len_of_non_continuous=len(continuous_check_bool)-len_of_continuous
                if  len_of_continuous==len(continuous_check_bool) \
		    or len_of_continuous>=10 \
		    or (len_of_continuous<10 and len_of_continuous>=3 and len_of_non_continuous<=2) \
                    or month_count[month_row_index]<=3\
                    or all(x == 0 for x in inv) :
		    #check the corresponding year
                    if max_match_year>0:
                        #st.write("max_match_year",max_match_year,"year_table",year_table)
                        PL_date_header=year_table.iloc[month_row_index,].apply(lambda x:str(int(x)))+\
                                                      month_table.iloc[month_row_index,].apply(lambda x:"" if x==0 else "0"+str(int(x)) if x<10 else str(int(x)))
                        
		        
                        if reporting_month not in list(PL_date_header):
                            #year_table.iloc[month_row_index,]=Fill_Year_To_Header(list(month_table.iloc[month_row_index,]),sheet_name,reporting_month)
                            PL_date_header=Fill_Year_To_Header(PL,month_row_index,list(month_table.iloc[month_row_index,]),sheet_name,reporting_month)
                            #st.write("PL_date_header2",PL_date_header)         
                    elif max_match_year==0:  # there is no year for all the months
		        #fill year to month
                        PL_date_header=Fill_Year_To_Header(PL,month_row_index,list(month_table.iloc[month_row_index,]),sheet_name,reporting_month)
                     
                        original_header=PL.iloc[month_row_index,]
                        PL_date_header_list=list(PL_date_header)
                   
                        d_str = ''
                        for i in range(len(PL_date_header)):
                            if PL_date_header[i]==0 or PL_date_header[i]=="0":
                                continue
                            else:
                                date=str(PL_date_header[i][4:6])+"/"+str(PL_date_header[i][0:4])
                                d_str +=",  "+str(PL.iloc[month_row_index,i])+" — "+ date
                
                        st.warning("Fail to identify **'Year'** for the date header in sheet '{}'. Filled year as:".format(sheet_name))
                        st.markdown(d_str[1:])
                    count_reporting_month=list(PL_date_header).count(reporting_month)
                   
                    if count_reporting_month==0: # there is no reporting_month
                       continue
                    elif count_reporting_month>1:  # there are duplicated months (more than one same months in header)
                        keywords = ["ytd", "year to date", "year-to-date","year_to_date","prior period","period ending","consolidated"]
                        duplicate_rm_columns = PL.columns[PL_date_header == reporting_month].tolist()  # the column index for duplicated reporting months
                        #st.write("PL",PL)
                        for col_idx in duplicate_rm_columns:
    			    # Search for "YTD", "Year to date", or "year-to_date"
                            if any(str(PL.iloc[row, col_idx]).strip().lower() in keywords for row in range(first_tenant_account_row)):
    			        # Change the corresponding value in `PL_date_header` to 0
                                PL_date_header[col_idx] = "0"
                        duplicate_rm_columns = PL.columns[PL_date_header == reporting_month].tolist()
                        if len(duplicate_rm_columns)==1:
                            return PL_date_header,month_row_index,PL.iloc[month_row_index,:]      
                        elif len(duplicate_rm_columns)>1:
                            # Compare the data below the month_row_index for these columns
                            for i, col_idx in enumerate(duplicate_rm_columns):
                                if i > 0:# Skip the first column since it's the one we're comparing others to
                                    # Extract values below month_row_index  
                                    values_below = PL_temp[col_idx].iloc[month_row_index + 1:].values        
                                    # Compare the values in this column with the first matching column
                               
                                    first_col_values_below = PL_temp[duplicate_rm_columns[0]].iloc[month_row_index + 1:].values  
                                    if (values_below == first_col_values_below).all():
                                        # If the values are the same, set the value of the current column in month_row_index to 0
                                        PL_date_header[col_idx] = "0"
                                    else:
                                        # If values are different
                                        st.error("There are more than one '{}/{}' header in sheet '{}'. Only one is allowed to identify the data column of '{}/{}'".\
			                         format(reporting_month[4:6],reporting_month[0:4],sheet_name,reporting_month[4:6],reporting_month[0:4]))
                                        st.stop()
                            if list(PL_date_header).count(reporting_month)==1:
                                return PL_date_header,month_row_index,PL.iloc[month_row_index,:]	
                    elif count_reporting_month==1:  # there is only one reporting month in the header
                        return PL_date_header,month_row_index,PL.iloc[month_row_index,:]	
                else:
                    continue
			
        
            # only one month in header, all the rows that have multiple months were out
            elif month_count[month_row_index]==1:
                col_month = next((col_no for col_no, val_month in enumerate(month_table.iloc[month_row_index, :]) if val_month != 0), 0)
                if month_table.iloc[month_row_index,col_month]!=int(reporting_month[4:]):
                    continue
		#if there are two month headers in the same column, conintue	
                if candidate_date!=[] and any(col_month == candidate_date_i[-1] for candidate_date_i in candidate_date):
                    continue
			
                PL_date_header= [reporting_month if x!=0 else 0 for x in month_table.iloc[month_row_index,:]]
                candidate_date.append([PL_date_header,month_row_index,PL.iloc[month_row_index,:],col_month] )
                continue

    if len(candidate_date)>1:
        #st.write(",".join([sublist[-1]+1 for sublist in candidate_date]))
        st.error("We detected {} month headers in sheet——'{}'. Please ensure there's only one month header for the data column.".format(len(candidate_date),sheet_name))
        st.stop()
    elif len(candidate_date)==1:	    
        return candidate_date[0][0:3]

    # there is no month/year in PL
    elif len(candidate_date)==0: 
	#  more than one column contain numeric data without any month date header
        if len(valid_col_index) > 1: 
            # search "current month" as reporting month
            current_month_cols=[]

            for col_i in valid_col_index:
                column = PL.iloc[0:first_tenant_account_row, col_i].reset_index(drop=True)
                if column.astype(str).str.contains('current month|current period|mtd|current', case=False, na=False).any():
                    current_month_cols.append(col_i)
                    current_month_rows = column.index[column.astype(str).str.contains(r'(current month|current period|mtd|current)', case=False, na=False)][0]
                elif sheet_type=="Sheet_Name_Occupancy" and column.astype(str).str.contains('#\\s*of\\s*days|total', case=False, na=False).any():
                    current_month_cols.append(col_i)
                    current_month_rows = column.index[column.astype(str).str.contains('#\\s*of\\s*days|total', case=False, na=False)][0]  
            if len(current_month_cols)==1:
                PL_date_header = [0] * PL_col_size
                PL_date_header[current_month_cols[0]] = reporting_month
                return PL_date_header,current_month_rows,PL.iloc[current_month_rows,:]
              
            # Didn't find key word "current month", remove the key word "ytd"...
            keywords = ["ytd", "year to date", "year-to-date","year_to_date","prior period","period ending","consolidated"]

            for col_idx in valid_col_index[:]:
    		# Search for "YTD", "Year to date", or "year-to_date"
                if any(str(PL.iloc[row, col_idx]).strip().lower() in keywords for row in range(first_tenant_account_row)):
                    # Change the corresponding value in `PL_date_header` to 0
                    valid_col_mask[col_idx] = False
            if np.sum(valid_col_mask) == 1:
                PL_date_header=[reporting_month if x else 0 for x in valid_col_mask]
                return PL_date_header,first_tenant_account_row-1,[]
              
        elif len(valid_col_index)==1:  #only one column contain numeric data
		
                only_numeric_column_value=PL_temp.iloc[:,valid_col_index[0]]
                # count the value in numeric column
                count_non = only_numeric_column_value.isna().sum()
                # Count string values
                count_str = only_numeric_column_value.apply(lambda x: isinstance(x, str)).sum()
                # Count numeric values
                count_num = only_numeric_column_value.apply(lambda x: pd.to_numeric(x, errors='coerce')).notna().sum()

                # numeric data is supposed to be more than character data
                if (count_str>0 and (count_num/count_str)<0.7):
                    st.error("Failed to identify Year/Month header for sheet: '{}', please add the month/year header and re-upload.".format(sheet_name))
                    st.stop()
                else:
		    # first_tenant_account_row -1 is the header row No. (used to remove the above rows, and prevent map new accounts)
                    #st.write("first_tenant_account_row",first_tenant_account_row,PL.iloc[first_tenant_account_row,:])
                    PL_date_header=[reporting_month if x else 0 for x in valid_col_mask]
                    return PL_date_header,first_tenant_account_row-1,[]

    st.error("Failed to identify {}/{} header for sheet: '{}', please add the month/year header and re-upload.".format(int(reporting_month[4:6]),reporting_month[0:4],sheet_name))
    st.stop()

# manage entity mapping in "Manage Mapping" 
def Manage_Entity_Mapping(operator):
    entity_mapping_updation=pd.DataFrame(\
	    columns=["Property_Name","Sheet_Name_Finance","Sheet_Name_Occupancy","Sheet_Name_Balance_Sheet","Column_Name"],\
            index=entity_mapping.index)
 
    entity_mapping_different_sheet_index= entity_mapping.index[(entity_mapping["DATE_SOLD_PAYOFF"]=="N") & ( entity_mapping["Finance_in_separate_sheets"]=="Y")]
   
    if len(entity_mapping_different_sheet_index)>0:
        with st.form(key="Property mapping"):
            col1,col2,col3,col4=st.columns([4,3,3,3])
            with col1:
                st.write("Property")
            with col2:
                st.write("P&L Sheetname")    
            with col3: 
                st.write("Census Sheetname")    
            with col4:
                st.write("BS Sheetname")  
  
            for entity_i in entity_mapping_different_sheet_index:
                col1,col2,col3,col4=st.columns([4,3,3,3])
                with col1:
                    st.write("")
                    st.write(entity_mapping.loc[entity_i,"Property_Name"])
                with col2:
                    new_value=st.text_input("new Financial sheet name",label_visibility="hidden",placeholder =entity_mapping.loc[entity_i,"Sheet_Name_Finance"],key="P&L"+entity_i)  
                    if new_value:
                        entity_mapping_updation.loc[entity_i,"Sheet_Name_Finance"]=new_value
                with col3: 
                    if not pd.isna(entity_mapping.loc[entity_i,"Sheet_Name_Occupancy"]):
                        new_value=st.text_input("New Occ sheet name",label_visibility="hidden",placeholder =entity_mapping.loc[entity_i,"Sheet_Name_Occupancy"],key="Census"+entity_i)
                        if new_value:
                            entity_mapping_updation.loc[entity_i,"Sheet_Name_Occupancy"]=new_value
                with col4:
                    if not pd.isna(entity_mapping.loc[entity_i,"Sheet_Name_Balance_Sheet"]):
                        new_value=st.text_input("New BS sheet name",label_visibility="hidden",placeholder =entity_mapping.loc[entity_i,"Sheet_Name_Balance_Sheet"],key="BS"+entity_i) 
                        if new_value:
                            entity_mapping_updation.loc[entity_i,"Sheet_Name_Balance_Sheet"]=new_value
            submitted = st.form_submit_button("Submit")
            if submitted:
                entity_mapping.update(entity_mapping_updation)
                Update_File_Onedrive(mapping_path,entity_mapping_filename,entity_mapping,operator,"CSV",None,entity_mapping_str_col)
                st.success("Updates mapping successfully!")
             
		
    entity_mapping_same_sheet_index= entity_mapping.index[(entity_mapping["DATE_SOLD_PAYOFF"]=="N")&(entity_mapping["Finance_in_separate_sheets"]=="N")]
    if len(entity_mapping_same_sheet_index)>0:
        with st.form(key="Mapping Property mapping"):
            col1,col2,col3,col4,col5=st.columns([4,3,3,3,4])
            with col1:
                st.write("Property")
            with col2:
                st.write("P&L Sheetname")    
            with col3: 
                st.write("Census Sheetname")    
            with col4:
                st.write("BS Sheetname") 
            with col5:
                st.write("Property name in header") 
  
            for entity_i in entity_mapping_same_sheet_index:
                col1,col2,col3,col4,col5=st.columns([4,3,3,3,4])
                with col1:
                    st.write("")
                    st.write(entity_mapping.loc[entity_i,"Property_Name"])
                with col2:
                    new_value=st.text_input("New Property name",label_visibility="hidden",placeholder =entity_mapping.loc[entity_i,"Sheet_Name_Finance"],key="PL"+entity_i)  
                    if new_value:
                        entity_mapping_updation.loc[entity_i,"Sheet_Name_Finance"]=new_value
                with col3: 
                    if not pd.isna(entity_mapping.loc[entity_i,"Sheet_Name_Occupancy"]):
                        new_value=st.text_input("New Occ sheet name",label_visibility="hidden",placeholder =entity_mapping.loc[entity_i,"Sheet_Name_Occupancy"],key="CS"+entity_i)  
                        if new_value:
                            entity_mapping_updation.loc[entity_i,"Sheet_Name_Occupancy"]=new_value
                with col4:
                    if not pd.isna(entity_mapping.loc[entity_i,"Sheet_Name_Balance_Sheet"]):
                        new_value=st.text_input("New BS sheet name",label_visibility="hidden",placeholder =entity_mapping.loc[entity_i,"Sheet_Name_Balance_Sheet"],key="BS"+entity_i) 
                        if new_value:
                            entity_mapping_updation.loc[entity_i,"Sheet_Name_Balance_Sheet"]=new_value
                with col5:
                    new_value=st.text_input("New column name",label_visibility="hidden",placeholder =entity_mapping.loc[entity_i,"Column_Name"],key="CN"+entity_i) 
                    if new_value:
                        entity_mapping_updation.loc[entity_i,"Column_Name"]=new_value
            submitted = st.form_submit_button("Submit")
            if submitted:
                entity_mapping.update(entity_mapping_updation)
                Update_File_Onedrive(mapping_path,entity_mapping_filename,entity_mapping,operator,"CSV",None,entity_mapping_str_col)
                st.success("Updates mapping successfully!")
        #st.write("new entity mapping",entity_mapping)
    return entity_mapping

# no cache 
def Manage_Account_Mapping(new_tenant_account_list,sheet_name="False",sheet_type_name="False"):
    global account_mapping
    st.warning("Please complete mapping for below new accounts:")
    count=len(new_tenant_account_list)
    Sabra_main_account_list=[np.nan] * count
    Sabra_second_account_list=[np.nan] * count
    Sabra_main_account_value=[np.nan] * count
    Sabra_second_account_value=[np.nan] * count
    form_key = f"form_{new_tenant_account_list[0]}_{sheet_name}_{sheet_type_name}"
    with st.form(key=form_key):  
        for i in range(count):
            if sheet_name=="False":
                st.markdown("## Map **'{}'** to Sabra account".format(new_tenant_account_list[i])) 
            elif sheet_type_name != "Occupancy":
                st.markdown("## Map **'{}'** （in {} sheet - '{}'） to Sabra account".format(new_tenant_account_list[i],sheet_type_name,sheet_name)) 
            else :   
                st.markdown("## Map **'{}'**（in census/hours sheet - '{}'）to Sabra account".format(new_tenant_account_list[i],sheet_name)) 
     
            col1,col2=st.columns(2) 
            with col1:
                st.write("Sabra main account")
                Sabra_main_account_list[i]=streamlit_tree_select.tree_select(parent_hierarchy_main,only_leaf_checkboxes=True,key=str(new_tenant_account_list[i])+"primary")  
            with col2:
                st.write("Sabra second account")
                Sabra_second_account_list[i]= streamlit_tree_select.tree_select(parent_hierarchy_second,only_leaf_checkboxes=True,key=str(new_tenant_account_list[i])+"second")
        
        #st.markdown('<p class="small-font">If you need to apply transformation (such as multiplying by -1) to some accounts, please email sli@sabrahealth.com.</p>', unsafe_allow_html=True)
        submitted = st.form_submit_button("Submit")    	    
            
        if submitted:
            for i in range(count):
                if len(Sabra_main_account_list[i]['checked'])==1:
                    Sabra_main_account_value[i]=Sabra_main_account_list[i]['checked'][0].upper()          
                elif len(Sabra_main_account_list[i]['checked'])>1:
                    if len(Sabra_main_account_list[i]['checked'])==2 and Sabra_main_account_list[i]['checked'][0]=="Management Fee":
                        Sabra_main_account_value[i]="T_MGMT_FEE"  
                    else:
                        st.warning("Only one to one mapping is allowed, but {} has more than one mappings.".format(new_tenant_account_list[i]))
                        st.stop()
                elif Sabra_main_account_list[i]['checked']==[] and Sabra_second_account_list[i]['checked']==[]:
                    st.warning("Please select Sabra account for '{}'".format(new_tenant_account_list[i]))
                    st.stop()
                elif Sabra_main_account_list[i]['checked']==[]:
                    Sabra_main_account_value[i]=''
            
                if Sabra_second_account_list[i]['checked']==[]:
                    Sabra_second_account_value[i]=''
                elif len(Sabra_second_account_list[i]['checked'])==1:
                    Sabra_second_account_value[i]=Sabra_second_account_list[i]['checked'][0].upper()
                elif len(Sabra_second_account_list[i]['checked'])>1:
                    st.warning("Only one to one mapping is allowed, but {} has more than one mappings.".format(new_tenant_account_list[i]))
                    st.stop()
        else:
            st.stop()
                
        #insert new record to the bottom line of account_mapping
        new_accounts_df = pd.DataFrame({'Sabra_Account': Sabra_main_account_value, 'Sabra_Second_Account': Sabra_second_account_value, 'Tenant_Account':list(map(lambda x:x.upper().strip(), new_tenant_account_list))})
        new_accounts_df["Operator"]=operator     
        new_accounts_df=new_accounts_df.merge(BPC_Account[["BPC_Account_Name","Category"]], left_on="Sabra_Account",right_on="BPC_Account_Name",how="left").drop(columns="BPC_Account_Name")  

        # if there are new revenue accounts,  check if revenue need multiply -1. 
        new_rev_accounts = new_accounts_df[new_accounts_df["Sabra_Account"].str.startswith("REV_")]

        if not new_rev_accounts.empty:
            original_revenue = account_mapping[account_mapping["Sabra_Account"].fillna("").str.startswith("REV_")]
            conversion_count = len(original_revenue[original_revenue["Conversion"].fillna("") == "*-1"])

            conversion_percentage=(conversion_count / len(original_revenue)) * 100 if len(original_revenue) > 0 else 0
            if conversion_percentage>=0.7:
                st.warning(f"Based on your previous revenue, below new revenue accounts will be adjusted by multiplying by -1. Please let us know at sli@sabrahealth.com if this process is incorrect.")
                st.warning(",".join(new_rev_accounts["Tenant_Account"]))
                new_accounts_df["Conversion"] = new_accounts_df["Sabra_Account"].apply(lambda x: "*-1" if x.startswith("REV_") else "")		
                st.session_state.email_body_for_Sabra+=f"""<p>New revenue accounts were added and adjusted by multiplying -1</p> """
               
            if conversion_percentage>0 and conversion_percentage<1:
                st.session_state.email_body_for_Sabra+=f"""<p>Not all the revenue accounts were adjusted by multiplying -1, please check.</p> """    

        # Create a dropdown for the last column
        account_mapping=pd.concat([account_mapping, new_accounts_df],ignore_index=True)

        if Update_File_Onedrive(mapping_path,account_mapping_filename,\
	             account_mapping[["Operator", "Sabra_Account", "Sabra_Second_Account", "Tenant_Account", "Conversion"]],\
		     operator,"XLSX",None,account_mapping_str_col):
            st.success("New accounts mapping were successfully saved.")   
        else:
            st.error("New accounts mapping were not successfully saved.") 
	
#@st.cache_data 
def Map_PL_Sabra(PL,entity,sheet_type,account_pool):
    # remove no need to map from account_mapping
    account_pool = account_pool[(account_pool["Sabra_Account"] != "NO NEED TO MAP") |
    (account_pool["Sabra_Second_Account"].notna())]
    #st.write(account_pool)
    
    main_account_mapping = account_pool.loc[account_pool["Sabra_Account"].apply(lambda x: pd.notna(x) and x.upper() != "NO NEED TO MAP")]
        # Concatenate main accounts with second accounts
    second_account_mapping = account_pool.loc[(pd.notna(account_pool["Sabra_Second_Account"])) & (account_pool["Sabra_Second_Account"] != "NO NEED TO MAP")]\
	[["Sabra_Second_Account","Tenant_Account", "Conversion"]]\
        .rename(columns={"Sabra_Second_Account": "Sabra_Account"})
    
    if second_account_mapping.shape[0]>0:
        second_account_mapping = second_account_mapping[second_account_mapping["Sabra_Account"].str.strip() != ""]
    
    # Ensure index name consistency
    PL.index.name = "Tenant_Account"
    PL = PL.reset_index(drop=False)
    
    # Filter main_account_mapping before the merge
    main_account_mapping_filtered = main_account_mapping[pd.notna(main_account_mapping["Sabra_Account"])][["Sabra_Account", "Tenant_Account", "Conversion"]] 
	
    PL = pd.concat([PL.merge(second_account_mapping, on="Tenant_Account", how="right"),\
                    PL.merge(main_account_mapping_filtered,   on="Tenant_Account", how="right")])

    #Remove blank or missing "Sabra_Account" values
    PL = PL[PL["Sabra_Account"].str.strip() != ""]
    PL.dropna(subset=["Sabra_Account"], inplace=True)
    # Conversion column
    PL = PL.reset_index(drop=True)
    conversion = PL["Conversion"].fillna(np.nan)
    #st.write("PL",PL) 
    if isinstance(entity, str):# one entity,  properties are in separate sheet
        month_cols=list(filter(lambda x:str(x[0:2])=="20",PL.columns))
        #Convert all values in the PL to numeric, coercing non-numeric values to NaN. Fill NaN values with 0.
        PL[month_cols] = PL[month_cols].apply(pd.to_numeric, errors='coerce').fillna(0)
	    
        for idx, conv in conversion.items():
            for month in month_cols:
                if pd.isna(conv):
                    continue
                elif conv == "/monthdays":
                    PL.loc[idx,month] /= monthrange(int(str(month)[0:4]), int(str(month)[4:6]))[1]
                elif conv == "*monthdays":
                    PL.loc[idx, month] *= monthrange(int(str(month)[0:4]), int(str(month)[4:6]))[1]
                elif conv.startswith("*"):


                    multiplier = float(conv.split("*")[1])
                    PL.loc[idx, month] *= multiplier
                else:
                    continue
        PL=PL.drop(["Conversion","Tenant_Account"], axis=1)
        PL["ENTITY"]=entity	    

    elif isinstance(entity, list):  # multiple properties are in one sheet,data's column name is "value" 
        monthdays=monthrange(int(str(reporting_month)[0:4]), int(str(reporting_month)[4:6]))[1]
        PL[entity] = PL[entity].apply(pd.to_numeric, errors='coerce').fillna(0)
        for idx, conv in conversion.items():
            if pd.isna(conv):
                continue
            elif conv == "/monthdays":
                PL.loc[idx, entity] /= monthdays
            elif conv == "*monthdays":
                PL.loc[idx, entity] *= monthdays
            elif conv.startswith("*"):
                multiplier = float(conv.split("*")[1])
                PL.loc[idx, entity] *= multiplier
            else:
                continue
 
        PL=PL.drop(["Conversion"], axis=1)
        PL = pd.melt(PL, id_vars=['Sabra_Account','Tenant_Account'], value_vars=entity, var_name='ENTITY')     
        PL=PL.drop(["Tenant_Account"], axis=1)
  
    # group by Sabra_Account
    PL = PL.groupby(by=['ENTITY',"Sabra_Account"], as_index=True).sum()
    PL= PL.apply(Format_Value)    # do these two step, so Total_PL can use combine.first 
    #st.write("PL",PL)  
    return PL   
	
@st.cache_data
def Compare_PL_Sabra(Total_PL,reporting_month):
#def Compare_PL_Sabra(Total_PL,PL_with_detail,reporting_month):
    month_list = [month for month in Total_PL.columns if month != reporting_month]
    rows = []
    for entity in entity_mapping.index:
        if entity not in Total_PL.index.get_level_values('ENTITY'):
            continue

        for timeid in month_list: 
	    # if this entity don't have data for this timeid(new/transferred property), skip to next month
            if Total_PL.loc[entity, timeid].isna().all():
                continue
            for matrix in BPC_Account.loc[BPC_Account["Category"]!="Balance Sheet","BPC_Account_Name"]: 
                try:
                    BPC_value=int(BPC_pull.loc[entity,matrix][timeid])
                except:
                    BPC_value=0
                try:
                    PL_value=int(Total_PL.loc[entity,matrix][timeid])
                except:
                    PL_value=0
                if BPC_value==0 and PL_value==0:
                    continue
                diff=BPC_value-PL_value
                diff_percent=abs(diff)/max(abs(PL_value),abs(BPC_value))
                if diff_percent>=0.001: 
                    new_row = {"TIME": timeid,"ENTITY": entity,"Sabra_Account": matrix,"Sabra": BPC_value, "P&L": PL_value,"Diff (Sabra-P&L)": diff,"Diff_Percent": diff_percent}
                    rows.append(new_row)
    diff_BPC_PL = pd.DataFrame(rows, columns=["TIME", "ENTITY", "Sabra_Account", "Sabra", "P&L", "Diff (Sabra-P&L)", "Diff_Percent"])
                   
    if diff_BPC_PL.shape[0]>0:
        #percent_discrepancy_accounts=diff_BPC_PL.shape[0]/(BPC_Account.shape[0]*len(Total_PL.columns))
        diff_BPC_PL=diff_BPC_PL.merge(BPC_Account[["Category","Sabra_Account_Full_Name","BPC_Account_Name"]],left_on="Sabra_Account",right_on="BPC_Account_Name",how="left")        
        diff_BPC_PL=diff_BPC_PL.merge(entity_mapping.reset_index(drop=False)[["ENTITY","Property_Name"]], on="ENTITY",how="left")
    return diff_BPC_PL
	
def color_missing(data):
    return f'background-color: rgb(255, 204, 204);'
def Compare_Total_with_Total(row1_PL,row2_Sabra,value_column,category,account_forluma):
    global email_body
    # Compute the difference (row1 - row2) for value_column
    diff = row1_PL[value_column].values - row2_Sabra[value_column].values
    # Create a new row for the difference
    #st.write("row1_PL",row1_PL,"row2_Sabra",row2_Sabra,"diff",diff)
    # Only keep values where abs(diff) > 10, else put np.nan
    diff_flat = diff.flatten()
    non_zero_indices = np.where(row1_PL[value_column].values.flatten() != 0)[0]

    # Identify indices where abs difference > 10 and total_xx is not 0(0 means there is no total account in P&L)
    significant_diff_indices = np.where(np.abs(diff_flat) > 10)[0]
    significant_diff_indices = np.intersect1d(significant_diff_indices, non_zero_indices)
    if len(significant_diff_indices) > 0:
        delta_properties_columns = [value_column[i] for i in significant_diff_indices]
        columns_to_keep=["Sabra_Account"] + delta_properties_columns 
        # Create filtered diff row
        diff_row = pd.DataFrame( data=[["Delta"] + [diff_flat[i] for i in significant_diff_indices]],columns=columns_to_keep)

        # Filter original rows to keep only selected columns
        row1_PL = row1_PL[columns_to_keep]
        row1_PL["Sabra_Account"] = row1_PL["Sabra_Account"] + f" ({account_forluma})"
        row2_Sabra = row2_Sabra[columns_to_keep]
        row2_Sabra["Sabra_Account"] = "Sabra report(calculated by mapping)"

        # Concatenate row1,row2, diff to create the final dataframe
        result_df = pd.concat([row2_Sabra,row1_PL,diff_row],ignore_index=True)
        st.error(f"The calculated {category} are inconsistent with those in the P&L. Please download the mapping and check it.")
        result_df=result_df.apply(Format_Value)
        result_df.rename(columns={"Sabra_Account": category.title()},inplace=True)
        #st.write(result_df)
        result_df_display = result_df.to_html(index=False)
        st.markdown(result_df_display, unsafe_allow_html=True)


        email_body+=f"<p>The calculated {category} are inconsistent with those in the P&L:</p>{result_df.to_html(index=False)}"
        
        return True
    return False

def Create_Account_Foluma(total_account):
    parts = []
    for _, row in total_account.iterrows():
        tenant = row["Tenant_Account"]
        sign = '-' if row["Conversion"] == "*-1" else '+'
        parts.append((sign, tenant))

    # Assemble the string, skip the sign on the first item if it's '+'
    expression = parts[0][1]  # First item without '+' sign
    for sign, tenant in parts[1:]:
        expression += f"{sign}{tenant}"

    return expression

def View_Summary(): 
    global Total_PL,reporting_month_data,placeholder,email_body
    def highlight_total(df):
        return ['color: blue']*len(df) if df.Sabra_Account.startswith("Total - ") or df.Sabra_Account.startswith("EBITDARM") else ''*len(df)
    Total_PL = Total_PL.fillna(0).infer_objects(copy=False)

    reporting_month_data=Total_PL[reporting_month].reset_index(drop=False)
    #st.write("reporting_month_data",reporting_month_data,reporting_month_data.index)
    reporting_month_data=reporting_month_data.merge(BPC_Account, left_on="Sabra_Account", right_on="BPC_Account_Name",how="left")	
    reporting_month_data=reporting_month_data.merge(entity_mapping[["Property_Name"]], on="ENTITY",how="left")
    #st.write("reporting_month_data",reporting_month_data,reporting_month_data.index)

    # check patient days ( available days > patient days)	
    
    check_patient_days = reporting_month_data[
    ((reporting_month_data["Sabra_Account"].str.startswith("A_")) | 
     (reporting_month_data["Category"] == "Patient Days")) &
    (~reporting_month_data["Sabra_Account"].str.startswith("TOTAL_", na=False))]
    check_patient_days.loc[check_patient_days['Category'] == 'Facility Information', 'Category'] = 'Operating Beds'
    check_patient_days=check_patient_days[["Property_Name","Category",reporting_month]].groupby(["Property_Name","Category"]).sum()
    check_patient_days = check_patient_days.fillna(0).infer_objects(copy=False)
    #check if available unit changed by previous month
    reporting_month_data,Total_PL,email_body=Check_Available_Units(reporting_month_data,Total_PL,check_patient_days,reporting_month,email_body)
   
    #check missing category ( example: total revenue= 0, total Opex=0...)	
    category_list=['Revenue','Patient Days','Operating Expenses',"Balance Sheet"]

    # Get unique entities
    entity_list = list(reporting_month_data["ENTITY"].unique())
	
    current_cagegory=reporting_month_data[["Property_Name","Category","ENTITY",reporting_month]][reporting_month_data["Category"].\
	    isin(category_list)].groupby(["Property_Name","Category","ENTITY"]).sum().reset_index(drop=False)

	
    full_category = pd.DataFrame(list(product(entity_list,category_list)), columns=['ENTITY', 'Category'])
    missing_category=full_category.merge(current_cagegory,on=['ENTITY', 'Category'],how="left")
    missing_category=missing_category[(missing_category[reporting_month]==0)|(missing_category[reporting_month].isnull())]
    missing_category[reporting_month]="Missing" 

    if missing_category.shape[0]>0:
        st.write("No data detected for below properties and accounts: ")
        missing_category=missing_category[["ENTITY",reporting_month,"Category"]].merge(entity_mapping[["Property_Name"]], on="ENTITY",how="left")[["Property_Name","Category",reporting_month]]
        missing_category=missing_category.rename(columns={'Property_Name':'Property',"Category":"Account category",reporting_month:reporting_month_display})
        st.dataframe(missing_category.style.map(color_missing, subset=[reporting_month_display]).hide(axis="index"))

        email_body+= f"<p> No data detected for below properties and accounts:</p>{missing_category.to_html(index=False)}"

    reporting_month_data =reporting_month_data.pivot_table(index=["Sabra_Account_Full_Name","Category"], columns="Property_Name", values=reporting_month,aggfunc='last')
    reporting_month_data.reset_index(drop=False,inplace=True)
    reporting_month_data.rename(columns={"Sabra_Account_Full_Name":"Sabra_Account"},inplace=True) 
    reporting_month_data=reporting_month_data.dropna(subset=["Sabra_Account"])
    sorter=["Facility Information","Patient Days","Revenue","Operating Expenses",\
	    "Non-Operating Expenses","Labor Expenses","Management Fee","Balance Sheet",\
	    "Additional Statistical Information","Government Funds","Total"]
    sorter=list(filter(lambda x:x in reporting_month_data["Category"].unique(),sorter))
	
    reporting_month_data.Category = reporting_month_data.Category.astype("category")
    reporting_month_data.Category = reporting_month_data.Category.cat.set_categories(sorter)
    reporting_month_data=reporting_month_data.sort_values(["Category"]) 
    reporting_month_data_temp = reporting_month_data[~reporting_month_data["Sabra_Account"].str.contains("in P&L", na=False)]
    reporting_month_data = pd.concat([reporting_month_data_temp.\
             groupby(by='Category', as_index=False,observed=False).\
	     sum().assign(Sabra_Account="Total_Sabra"), reporting_month_data]).\
	     sort_values(by='Category', kind='stable', ignore_index=True)[reporting_month_data.columns]

    set_empty=list(reporting_month_data.columns)
    set_empty.remove("Category")
    set_empty.remove("Sabra_Account")
    for i in range(reporting_month_data.shape[0]):
        if reporting_month_data.loc[i,"Sabra_Account"]=="Total_Sabra":
            reporting_month_data.loc[i,"Sabra_Account"]="Total - "+reporting_month_data.loc[i,'Category']
            if reporting_month_data.loc[i,'Category'] in ["Facility Information","Additional Statistical Information","Balance Sheet"]:                
                reporting_month_data.loc[i,set_empty]=np.nan

    reporting_month_data = reporting_month_data.set_index("Sabra_Account")
    reporting_month_data.drop(columns=["Category"], inplace=True)

    # Calculate DARM
    EBITDARM = reporting_month_data.loc["Total - Revenue"] - reporting_month_data.loc["Total - Operating Expenses"]

    # Convert it into a DataFrame row
    EBITDARM_row = pd.DataFrame([EBITDARM], index=["EBITDARM"])
    EBITDARM_row.index.name = "Sabra_Account"

    if "Total - Non-Operating Expenses" in reporting_month_data.index:
        # Insert above 'Total - Non-Operating Expenses'
        reporting_month_data = pd.concat([
            reporting_month_data.loc[:'Total - Non-Operating Expenses'].iloc[:-1],
            EBITDARM_row,
            reporting_month_data.loc['Total - Non-Operating Expenses':]])
    else:
        # Append to the bottom
        reporting_month_data = pd.concat([reporting_month_data, EBITDARM_row])


    # Reset index to have "Sabra_Account" as a column again
    entity_columns=reporting_month_data.columns
    reporting_month_data = reporting_month_data.reset_index(drop=False)
    reporting_month_data["Total"] = reporting_month_data[entity_columns].sum(axis=1)
    reporting_month_data=reporting_month_data[["Sabra_Account","Total"]+list(entity_columns)]
	
    PL_total_names=["Total Patient Days in P&L","Total Revenue in P&L","Total OPEX in P&L","Total Expense in P&L"]
    PL_total = reporting_month_data[reporting_month_data["Sabra_Account"].isin(PL_total_names)]
    missing_total_accounts_PL = PL_total[PL_total.drop(columns='Sabra_Account').eq(0).any(axis=1)]
    
    if not missing_total_accounts_PL.empty:
        st.session_state.email_body_for_Sabra += f"""
        <p>Below properties are missing total accounts:</p>
        {missing_total_accounts_PL.to_html(index=False, escape=False)}
    """
    PL_total = PL_total[PL_total["Total"] != 0]

    # DataFrame with all other rows
    PL_total = PL_total.drop(columns="Total")	    
    value_column=list(entity_columns)
    
    if PL_total.shape[0]>0:
        download_mapping=False
        compare_metric=PL_total["Sabra_Account"].tolist()
        if "Total Patient Days in P&L" in compare_metric:
            row1_PL = PL_total[PL_total["Sabra_Account"] == "Total Patient Days in P&L"]
            row2_Sabra = reporting_month_data[reporting_month_data["Sabra_Account"] == "Total - Patient Days"]
            total_account = account_mapping[account_mapping["Sabra_Account"] == "TOTAL_PD"]
            account_string=Create_Account_Foluma(total_account)
            if Compare_Total_with_Total(row1_PL,row2_Sabra,value_column,"total patient days",account_string):
                download_mapping=True

            
        if "Total Revenue in P&L" in compare_metric:
            row1_PL = PL_total[PL_total["Sabra_Account"] == "Total Revenue in P&L"]
            row2_Sabra = reporting_month_data[reporting_month_data["Sabra_Account"] == "Total - Revenue"]
            total_account = account_mapping[account_mapping["Sabra_Account"] == "TOTAL_REV"]
            account_string=Create_Account_Foluma(total_account)
            if Compare_Total_with_Total(row1_PL,row2_Sabra,value_column,"total revenue",account_string):
                download_mapping=True
        if "Total OPEX in P&L" in compare_metric:
            row1_PL = PL_total[PL_total["Sabra_Account"] == "Total OPEX in P&L"]
            row2_Sabra = reporting_month_data[reporting_month_data["Sabra_Account"] == "Total - Operating Expenses"]
            total_account = account_mapping[account_mapping["Sabra_Account"] == "TOTAL_OPEX"]
            account_string=Create_Account_Foluma(total_account)
            if Compare_Total_with_Total(row1_PL,row2_Sabra,value_column,"total operating expense",account_string):
                download_mapping=True
        if "Total Expense in P&L" in compare_metric:
            row1_PL = PL_total[PL_total["Sabra_Account"] == "Total Expense in P&L"]
            sabra_total_accounts = ["Total - Operating Expenses", "Total - Non-Operating Expenses", "Total - Management Fee"]
            row2_Sabra = reporting_month_data[reporting_month_data["Sabra_Account"].isin(sabra_total_accounts)]

            # Sum the numeric columns across the filtered rows
            row2_Sabra = row2_Sabra.drop(columns=["Sabra_Account"]).sum().to_frame().T
            row2_Sabra.insert(0, "Sabra_Account", "Total - Expense")
            total_account = account_mapping[account_mapping["Sabra_Account"] == "TOTAL_EXP"]
            account_string=Create_Account_Foluma(total_account)
            if Compare_Total_with_Total(row1_PL,row2_Sabra,value_column,"total expense",account_string):
                download_mapping=True
            
        if download_mapping:
            download_report(account_mapping,"mapping to check inconsistence")

    reporting_month_data = reporting_month_data[~reporting_month_data["Sabra_Account"].isin(PL_total_names+["Total - Total"])]

    placeholder = st.empty()
    placeholder.markdown("""
            <div style="background-color: #fff1ad; padding: 10px; border-radius: 5px;">
            ⚠️ <b style="font-size: 18px;">Reminder:</b> Please make sure click the 
            '<b style="font-size: 18px;">Confirm and Upload</b>' button at the bottom of the report 
            to complete the upload!
            </div>
            """,unsafe_allow_html=True)
    with st.expander("{} {} reporting".format(operator,reporting_month_display) ,expanded=True):
        ChangeWidgetFontSize("{} {} reporting".format(operator,reporting_month_display), '25px')
        col1,col2=st.columns([1,5])
        with col1:
            download_report(reporting_month_data,"Report")
        with col2:
            download_report(account_mapping,"accounts mapping")
        reporting_month_data=reporting_month_data.apply(Format_Value)
        reporting_month_data=reporting_month_data.fillna(0).infer_objects(copy=False)
        reporting_month_data=reporting_month_data.replace(0,'')
        styled_table = (reporting_month_data.style.set_table_styles(styles).apply(highlight_total, axis=1).format(precision=0, thousands=",").hide(axis="index").to_html(escape=False)) # Use escape=False to allow HTML tags
        # Display the HTML using st.markdown
        st.markdown(styled_table, unsafe_allow_html=True)
        st.write("")
        if len(reporting_month_data.columns)>3:
            show_column=["Sabra_Account", "Total"]
        else:
            show_column=["Sabra_Account"]
        

        summary_for_email = reporting_month_data[
                          (reporting_month_data["Sabra_Account"].isin([
                          "Total - Patient Days", 
                          "Total - Revenue", 
                          "Total - Operating Expenses", 
                          "Total - Non-Operating Expenses",
			  "EBITDARM",
                          "Total - Management Fee"  
                          ])) |
                          ((reporting_month_data["Sabra_Account"].str.startswith("Operating Beds-")) & 
                          (reporting_month_data["Total"].notna()) &  # Ensures "Total" is not NaN
                          (reporting_month_data["Total"] != 0) & 
                          (reporting_month_data["Total"] != ""))
                          ][show_column + list(entity_columns)].copy()  # Create a copy to modify data safely

        # Remove "Total - " from the Sabra_Account column
        summary_for_email["Sabra_Account"] = summary_for_email["Sabra_Account"].str.replace(r"^Total - ", "", regex=True)
        summary_for_email.rename(columns={"Sabra_Account": reporting_month_display }, inplace=True)

        #st.write("reporting_month_data",reporting_month_data,"summary_for_email",summary_for_email)	
        summary_for_email.columns.name = None 
        total_email_body=f"<p>Here is the summary for your reference:</p>{summary_for_email.to_html(index=False)}"+email_body
        return total_email_body
# no cache
def Submit_Upload(total_email_body,SHAREPOINT_FOLDER):
    upload_reporting_month=Total_PL[reporting_month].reset_index(drop=False)
    upload_reporting_month = upload_reporting_month[~upload_reporting_month['Sabra_Account'].str.startswith("TOTAL_")]
    upload_reporting_month["TIME"]=reporting_month
    upload_reporting_month=upload_reporting_month.rename(columns={reporting_month:"Amount"})
    current_time = datetime.now(pytz.timezone('America/Los_Angeles')).strftime("%H:%M")
    upload_reporting_month["Latest_Upload_Time"]=str(today)+" "+current_time
    upload_reporting_month["Operator"]=operator
    upload_reporting_month=upload_reporting_month.apply(Format_Value)

    if Update_File_Onedrive(master_template_path,monthly_reporting_filename,upload_reporting_month,operator,"CSV",None,None):
        st.success("{} {} reporting data was uploaded successfully!".format(operator,reporting_month[4:6]+"/"+reporting_month[0:4]))
        placeholder.empty()
    else: 
        st.write(" ")  #----------record into error report------------------------	
        # save original tenant P&L to OneDrive

    subject = "Confirmation of {} {} reporting".format(operator,reporting_month_display)
    # Get 'Asset_Manager' from entity_mapping

    unique_asset_managers = (entity_mapping['Asset_Manager'].dropna().str.split(',').explode().str.strip().dropna().unique())
    unique_asset_managers = unique_asset_managers.tolist()
    receiver_email_list = operator_email.split(",") + ["twarner@sabrahealth.com","sli@sabrahealth.com","jmanalastas@sabrahealth.com","amallawa@sabrahealth.com"]
    
    if '@*' in operator_email:
        st.error("Please update email address (in 'Menu' - 'Edit Account') to ensure you receive confirmation email.")
    # Append these unique values to receiver_list
    receiver_email_list.extend(unique_asset_managers)
  
    # Send the confirmation email
    format_total_email_body= f"""
    <html>
    <body>
        <p>Dear {operator} team,</p>
	<p>Thanks for submitting {operator} {reporting_month_display} reporting data.</p>"""+total_email_body+f"""<p>Best regards,</p>
        <p>Sabra Healthcare REIT.</p>
    </body>
    </html>"""

    if not st.session_state.email_sent:
        receiver_email_list=unique_asset_managers #["sli@sabrahealth.com"]   
        Send_Confirmation_Email(receiver_email_list, subject, format_total_email_body) 
        
        if email_body!="" or st.session_state.email_body_for_Sabra!="":
            Send_Confirmation_Email(["sli@sabrahealth.com"], "!!! Issues for {} {} reporting".format(operator,reporting_month_display), email_body+st.session_state.email_body_for_Sabra)    
        st.session_state.email_sent = True
def Check_Sheet_Name_List(uploaded_file,sheet_type):
    global PL_sheet_list
    try:
        wb = load_workbook(uploaded_file, data_only=True)
        PL_sheet_list = wb.sheetnames        
    except TypeError as e:
        # Check if the specific TypeError message matches
        error_message = str(e)
        if "<class 'openpyxl.styles.named_styles._NamedCellStyle'>.name should be <class 'str'> but value is <class 'NoneType'>" in error_message:
            st.write("Error: The Excel file is corrupted or has invalid styles. Please open the file and re-save it, which sometimes resolves such issues.")
        else:
            st.write("Fail to get the sheet names in {}. Please save the file as 'xlsx' and re-upload.".format(sheet_type))
    except Exception as e:
        st.write("Fail to get the sheet names in {}. Please save the file as 'xlsx' and re-upload.".format(sheet_type))

    if sheet_type=="Finance":
        missing_PL_sheet_property = entity_mapping[(~entity_mapping["Sheet_Name_Finance"].isin(PL_sheet_list))|(pd.isna(entity_mapping["Sheet_Name_Finance"]))]
        missing_PL_sheet_property_Y=missing_PL_sheet_property.loc[missing_PL_sheet_property["Finance_in_separate_sheets"]=="Y",:]
        missing_PL_sheet_property_N=missing_PL_sheet_property.loc[missing_PL_sheet_property["Finance_in_separate_sheets"]=="N",:]
        missing_occ_sheet_property = entity_mapping[(entity_mapping["Sheet_Name_Occupancy"].isin(PL_sheet_list)==False) & (pd.notna(entity_mapping["Sheet_Name_Occupancy"]))& (entity_mapping["Sheet_Name_Finance"] != entity_mapping["Sheet_Name_Occupancy"])]
        missing_occ_sheet_property_Y=missing_occ_sheet_property.loc[missing_occ_sheet_property["Finance_in_separate_sheets"]=="Y",:]
        missing_occ_sheet_property_N=missing_occ_sheet_property.loc[missing_occ_sheet_property["Finance_in_separate_sheets"]=="N",:]
        missing_BS_sheet_property = entity_mapping[(entity_mapping["BS_separate_excel"]=="N") &(pd.notna(entity_mapping["Sheet_Name_Balance_Sheet"]))& (entity_mapping["Sheet_Name_Finance"] != entity_mapping["Sheet_Name_Balance_Sheet"])&(entity_mapping["Sheet_Name_Balance_Sheet"].isin(PL_sheet_list)==False)]		
        missing_BS_sheet_property_Y=missing_BS_sheet_property.loc[missing_BS_sheet_property["Finance_in_separate_sheets"]=="Y",:]
        missing_BS_sheet_property_N=missing_BS_sheet_property.loc[missing_BS_sheet_property["Finance_in_separate_sheets"]=="N",:]    
        total_missing_Y=missing_PL_sheet_property_Y.shape[0]+missing_occ_sheet_property_Y.shape[0]+missing_BS_sheet_property_Y.shape[0]
        total_missing_N=missing_PL_sheet_property_N.shape[0]+missing_occ_sheet_property_N.shape[0]+missing_BS_sheet_property_N.shape[0]
    elif sheet_type=="BS": # BS in another excel file
        missing_BS_sheet_property = entity_mapping[((entity_mapping["BS_separate_excel"]=="Y") & (entity_mapping["Sheet_Name_Balance_Sheet"].isin(PL_sheet_list)==False))|(pd.isna(entity_mapping["Sheet_Name_Balance_Sheet"]))]
        missing_BS_sheet_property_Y=missing_BS_sheet_property.loc[missing_BS_sheet_property["Finance_in_separate_sheets"]=="Y",:]
        missing_BS_sheet_property_N=missing_BS_sheet_property.loc[missing_BS_sheet_property["Finance_in_separate_sheets"]=="N",:]  
        total_missing_Y=missing_BS_sheet_property_Y.shape[0]
        total_missing_N=missing_BS_sheet_property_N.shape[0]

    if total_missing_Y+total_missing_N==0:        
        return entity_mapping,wb
    

    with st.form(key=sheet_type):
        if sheet_type=="Finance":
            if  total_missing_Y>0:
                if missing_PL_sheet_property_Y.shape[0]>0:
                    for entity_i in missing_PL_sheet_property_Y.index:
                        st.warning("Please provide P&L sheet name for {}".format(entity_mapping.loc[entity_i,"Property_Name"]))
                        missing_PL_sheet_property_Y.loc[entity_i,"Sheet_Name_Finance"]=st.selectbox("Original P&L sheet name: {}".format(entity_mapping.loc[entity_i,"Sheet_Name_Finance"]),[""]+PL_sheet_list,key=entity_i+"PL_Y")
                if missing_occ_sheet_property_Y.shape[0]>0:
                    for entity_i in missing_occ_sheet_property_Y.index:
                        st.warning("Please provide Census sheet name for {}".format(entity_mapping.loc[entity_i,"Property_Name"]))
                        missing_occ_sheet_property_Y.loc[entity_i,"Sheet_Name_Occupancy"]=st.selectbox("Original Census sheet name: {}".format(entity_mapping.loc[entity_i,"Sheet_Name_Occupancy"]),[""]+PL_sheet_list,key=entity_i+"occ_Y")
            if  total_missing_N>0:
                if missing_PL_sheet_property_N.shape[0]>0:
                    st.warning("Please provide P&L sheet name for below properties:")
                    st.dataframe(missing_PL_sheet_property_N[["Property_Name"]],hide_index=True)
                    PL_sheet=st.selectbox("Original P&L sheet name: {}".format(entity_mapping.loc[missing_PL_sheet_property_N.index[0],"Sheet_Name_Finance"]),[""]+PL_sheet_list,key="P&L_N")
                if missing_occ_sheet_property_N.shape[0]>0:
                    st.warning("Please provide Census sheet name for below properties:")
                    st.dataframe(missing_occ_sheet_property_N[["Property_Name"]],hide_index=True)
                    occ_sheet=st.selectbox("Original Census sheet name: {}".format(entity_mapping.loc[missing_occ_sheet_property_N.index[0],"Sheet_Name_Occupancy"]),[""]+PL_sheet_list,key="occ_N")


        if missing_BS_sheet_property_Y.shape[0]>0:
            for entity_i in missing_BS_sheet_property_Y.index:
                st.warning("Please provide Balance Sheet sheet name for {}".format(entity_mapping.loc[entity_i,"Property_Name"]))
                missing_BS_sheet_property_Y.loc[entity_i,"Sheet_Name_Balance_Sheet"]=st.selectbox("Original Balance Sheet name: {}".format(entity_mapping.loc[entity_i,"Sheet_Name_Balance_Sheet"]),[""]+PL_sheet_list,key=entity_i+"bs_Y")   
            
        if missing_BS_sheet_property_N.shape[0]>0:
            st.warning("Please provide Balance sheet name for for below properties:")
            st.dataframe(missing_BS_sheet_property_N[["Property_Name"]],hide_index=True)
            BS_sheet=st.selectbox("Original Balance sheet name: {}".format(entity_mapping.loc[missing_BS_sheet_property_N.index[0],"Sheet_Name_Balance_Sheet"]),[""]+PL_sheet_list,key="BS_N")         
            
        submitted = st.form_submit_button("Submit")
           
    if submitted:
            if sheet_type=="Finance":
                if (missing_PL_sheet_property_Y.shape[0]>0 and missing_PL_sheet_property_Y["Sheet_Name_Finance"].isna().any()) or (missing_occ_sheet_property_Y.shape[0]>0 and missing_occ_sheet_property_Y["Sheet_Name_Occupancy"].isna().any()) or (missing_BS_sheet_property_Y.shape[0]>0 and missing_BS_sheet_property_Y["Sheet_Name_Balance_Sheet"].isna().any()):
                    st.error("Please complete above mapping.")
                    st.stop()
                else:
                    if missing_PL_sheet_property_Y.shape[0]>0:
	                # each property in seperate sheet, so the sheet names should be unique
                        duplicates = missing_PL_sheet_property_Y[missing_PL_sheet_property_Y.duplicated('Sheet_Name_Finance', keep=False)]
                        # Group by 'Sheet_Name_Finance' and get corresponding 'Property Name'
                        if not duplicates.empty:
                            grouped = duplicates.groupby("Sheet_Name_Finance")["Property_Name"].apply(lambda x: ', '.join(x)).reset_index()
                            for _, row in grouped.iterrows():
                                st.error(f"The sheet names are supposed to be different for {row['Property_Name']} .")
                                st.stop()
                        for entity_i in missing_PL_sheet_property_Y.index: 
                            new_finance_sheet_name=missing_PL_sheet_property_Y.loc[entity_i,"Sheet_Name_Finance"]
                            if new_finance_sheet_name in entity_mapping["Sheet_Name_Finance"].values:
                                property = entity_mapping.loc[entity_mapping["Sheet_Name_Finance"] == new_finance_sheet_name, "Property_Name"].iloc[0]
                                st.error(f"The '{new_finance_sheet_name}' sheet is for {property}, please choose another sheet for {missing_PL_sheet_property_Y.loc[entity_i, 'Property_Name']}")
                                st.stop()
                            entity_mapping.loc[entity_i,"Sheet_Name_Finance"]=new_finance_sheet_name
	
                    if missing_occ_sheet_property_Y.shape[0]>0:
                        for entity_i in missing_occ_sheet_property_Y.index:
                            entity_mapping.loc[entity_i,"Sheet_Name_Occupancy"]=missing_occ_sheet_property_Y.loc[entity_i,"Sheet_Name_Occupancy"]
                    if missing_BS_sheet_property_Y.shape[0]>0:
                        for entity_i in missing_BS_sheet_property_Y.index:
                            entity_mapping.loc[entity_i,"Sheet_Name_Balance_Sheet"]=missing_BS_sheet_property_Y.loc[entity_i,"Sheet_Name_Balance_Sheet"]

                if (missing_PL_sheet_property_N.shape[0]>0 and PL_sheet== "")\
			or (missing_occ_sheet_property_N.shape[0]>0 and occ_sheet== "") \
			or (missing_BS_sheet_property_N.shape[0]>0 and BS_sheet== ""):
                        st.error("Please complete above mapping.")
                        st.stop()
                else:
                    if missing_PL_sheet_property_N.shape[0]>0:
                        entity_mapping.loc[missing_PL_sheet_property_N.index,"Sheet_Name_Finance"]=PL_sheet
                    if missing_occ_sheet_property_N.shape[0]>0:
                        entity_mapping.loc[missing_occ_sheet_property_N.index,"Sheet_Name_Occupancy"]=occ_sheet
                    if missing_BS_sheet_property_N.shape[0]>0:
                        entity_mapping.loc[missing_BS_sheet_property_N.index,"Sheet_Name_Balance_Sheet"]=BS_sheet

            elif sheet_type=="BS":
                if (missing_BS_sheet_property_Y.shape[0]>0 and missing_BS_sheet_property_Y["Sheet_Name_Balance_Sheet"].isna().any()):
                    st.error("Please complete Balance Sheet mapping.")
                    st.stop()
                else:
                    if missing_BS_sheet_property_Y.shape[0]>0:
                        for entity_i in missing_BS_sheet_property_Y.index:
                            entity_mapping.loc[entity_i,"Sheet_Name_Balance_Sheet"]=missing_BS_sheet_property_Y.loc[entity_i,"Sheet_Name_Balance_Sheet"]

                if missing_BS_sheet_property_N.shape[0]>0:
                    if BS_sheet== "":
                        st.error("Please complete Balance Sheet mapping.")
                        st.stop()
                    else:
                        entity_mapping.loc[missing_BS_sheet_property_N.index,"Sheet_Name_Balance_Sheet"]=BS_sheet
    else:
        st.stop()
                
             
    # update entity_mapping in onedrive 
    Update_File_Onedrive(mapping_path,entity_mapping_filename,entity_mapping,operator,"CSV",None,entity_mapping_str_col)
    st.success("Mapping updated. Please keep the sheet names consistent to avoid remapping them each time. ")
    return entity_mapping,wb

# don't use cache
def View_Discrepancy(): 
    global diff_BPC_PL	
    if diff_BPC_PL.shape[0]>0:
        # save all the discrepancy 
        diff_BPC_PL["Operator"]=operator
        diff_BPC_PL=diff_BPC_PL.merge(entity_mapping[["GEOGRAPHY","LEASE_NAME","FACILITY_TYPE","INV_TYPE"]],on="ENTITY",how="left")
	# insert dims to diff_BPC_PL
        diff_BPC_PL["TIME"]=diff_BPC_PL["TIME"].apply(lambda x: "{}.{}".format(str(x)[0:4],month_abbr[int(str(x)[4:6])]))

	# only display the big discrepancy
        edited_diff_BPC_PL=diff_BPC_PL[diff_BPC_PL["Diff_Percent"]>0.15] 
        if edited_diff_BPC_PL.shape[0]>0:
            st.error("Below P&L data doesn't tie to Sabra data.  Please leave comments for discrepancy in below table.")
            edited_diff_BPC_PL.loc[:, "Type comments below"] = ""
            edited_diff_BPC_PL = st.data_editor(
	    edited_diff_BPC_PL,
	    width = 1200,
	    column_order=("Property_Name","TIME","Category","Sabra_Account_Full_Name","Sabra","P&L","Diff (Sabra-P&L)","Type comments below"),
	    hide_index=True,
	    disabled=("Property_Name","TIME","Category","Sabra_Account_Full_Name","Sabra","P&L","Diff (Sabra-P&L)"),
	    column_config={
       		"Sabra_Account_Full_Name": "Sabra_Account",
       		 "Property_Name": "Property",
		 "TIME":"Month",
		"P&L":st.column_config.TextColumn(
			"Tenant P&L",help="Tenant P&L is aggregated by detail tenant accounts connected with 'Sabra Account'"),
        	"Diff (Sabra-P&L)": st.column_config.TextColumn(
            		"Diff (Sabra-P&L)",help="Diff = Sabra-TenantP&L"),
		"Sabra": st.column_config.TextColumn(
            		"Sabra",help="Sabra data for previous month"),
		 "Type comments below":st.column_config.TextColumn(
            		"Type comments below",
            		help="Please provide an explanation and solution on discrepancy, like: confirm the changed. overwrite Sabra data with new one...",
			disabled =False,
            		required =False)
		}) 

            col1,col2=st.columns([1,4]) 
            with col1:
                submit_com=st.button("Submit comments")

            if submit_com:
                with st.empty():
                    with col2:
                        st.markdown("✔️ :green[Comments uploaded]")
                        st.write(" ")
                    # insert comments to diff_BPC_PL
                    diff_BPC_PL=pd.merge(diff_BPC_PL,edited_diff_BPC_PL[["Property_Name","TIME","Sabra_Account_Full_Name","Type comments below"]],on=["Property_Name","TIME","Sabra_Account_Full_Name"],how="left")
                    # save discrepancy data to OneDrive
                    if len(Total_PL.columns)>1 and diff_BPC_PL.shape[0]>0:
                        download_report(diff_BPC_PL[["Property_Name","TIME","Category","Sabra_Account_Full_Name","Sabra","P&L","Diff (Sabra-P&L)"]],"discrepancy")
                        Update_File_Onedrive(master_template_path,discrepancy_filename,diff_BPC_PL,operator,"CSV",None,None)
        
        else:
            st.success("All previous data in P&L ties with Sabra data")
    else:
            st.success("All previous data in P&L ties with Sabra data")



def Is_Reporting_Month(single_string):
    month=reporting_month[4:6]
    year=reporting_month[0:4]
    if single_string!=single_string or pd.isna(single_string):
        return False
    if isinstance(single_string, datetime) and int(single_string.month)==int(month):
        return True
    if isinstance(single_string, (int,float)):
        return False
    single_string=str(single_string).lower()
    if any([month_i in single_string for month_i in month_dic_word[int(month)]]):
        return True
    if (year in single_string) or (year[2:4] in single_string):
        single_string=single_string.replace(year,"").replace(year[2:4],"").replace("30","").replace("31","").replace("29","").replace("28","").\
	              replace("/","").replace("-","").replace(" ","").replace("_","").replace("asof","").replace("actual","").replace("mtd","")

        if str(int(month)) in single_string: 
            return True
    return False

def Identify_Column_Name_Header(PL,tenant_account_col_values,entity_list,sheet_name): 
    entity_without_propertynamefinance = entity_mapping[ (entity_mapping.index.isin(entity_list)) & \
    ((entity_mapping['Column_Name'].isna()) | (entity_mapping['Column_Name'].str.strip() == ""))].index.tolist()
    column_name_list_in_mapping=[str(x).upper().strip() for x in entity_mapping.loc[entity_list]["Column_Name"] if pd.notna(x) and str(x).strip()]
    max_match=[]
    #st.write("tenant_account_col_values",tenant_account_col_values)	
    # Create a set of tenant accounts that need mapping
    accounts_to_map = {account for account, sabra_account in zip(account_mapping['Tenant_Account'], account_mapping['Sabra_Account']) if sabra_account!= 'NO NEED TO MAP'}

    # Create a boolean mask using a list comprehension
    tenant_account_row_mask = [account in accounts_to_map for account in tenant_account_col_values]
    #st.write("tenant_account_row_mask",tenant_account_row_mask)	
    #first_tenant_account_row is the row number for the first tenant account (except for no need to map)
    first_tenant_account_row=tenant_account_row_mask.index(max(tenant_account_row_mask))
    #st.write("first_tenant_account_row",first_tenant_account_row)
    month_mask=[]
    # search the row with property column names	
    for row_i in range(first_tenant_account_row):
        canditate_row = list(map(lambda x: str(int(x)).upper().strip() if pd.notna(x) and isinstance(x, float) else str(x).upper().strip(), list(PL.iloc[row_i, :])))
        #st.write("canditate_row",canditate_row)
        match_names = [item for item in canditate_row if item in column_name_list_in_mapping] 
        #st.write("match_names",match_names)
	# find the property name header row, transferred them into entity id
        if len(match_names)>0 and sorted(match_names)==sorted(column_name_list_in_mapping) and len(entity_without_propertynamefinance)==0: 
           # property name column header is unique and match with entity mapping
            mapping_dict = {column_name_list_in_mapping[i]: entity_list[i] for i in range(len(column_name_list_in_mapping))}
            mapped_entity = [mapping_dict[property] if property in mapping_dict else "0" for property in canditate_row]
            return row_i,mapped_entity
	
        elif len(match_names)>len(max_match):
            max_match=match_names
            header_row=canditate_row
            max_match_row=row_i
	    # find the column name row, but it has a different length with entity_mapping's "column name"
            if len(match_names)==len(column_name_list_in_mapping):
                break
        if len(max_match)>2:
            break
    if len(max_match)==0: # there is no any column name header at all
        st.error("Fail to identify facility name header in sheet '{}'. The previous header names are as below. Please add and re-upload.".format(sheet_name))
        st.write(',    '.join(column_name_list_in_mapping))
        st.stop()
    elif len(max_match)>0: # only part of entities have column name in P&L  
        rest_column_names=[str(x) for x in PL.iloc[max_match_row,:] if pd.notna(x) and str(x).upper().strip() not in column_name_list_in_mapping]
        duplicate_check = [name for name in set(max_match) if max_match.count(name) > 1]
        if len(duplicate_check)>0:
	    # there may has more than one month for each property, only find the column of reporting month
            # Check reporting month above first_tenant_account_row
            mask_table = PL.iloc[0:first_tenant_account_row, :].apply(lambda col: col.map(Is_Reporting_Month))

            month_counts=pd.Series(np.sum(mask_table.values, axis=1))		
            if all(month_count==0 for month_count in month_counts): # there is no month
                st.error("Detected duplicated column names—— {} in sheet '{}'. Please fix and re-upload.".format(", ".join(f"'{item}'" for item in duplicate_check),sheet_name))
                st.stop()
            # month_row_index is the row having most reporting month
            max_month_index = month_counts.idxmax()
            month_mask = mask_table.iloc[max_month_index,:]
            filter_header_row =[item if item in column_name_list_in_mapping else 0 for item in header_row]
            filter_header_row = [item if is_month else 0 for item, is_month in zip(filter_header_row, month_mask)]

            duplicate_check=[item for item in set(filter_header_row) if filter_header_row.count(item) > 1 and item!=0]		
            # after apply month_mask, the column_name match with that in entity_mapping		
            if len(duplicate_check)==0 and sorted([x for x in filter_header_row if x != 0]) == sorted(column_name_list_in_mapping) and len(entity_without_propertynamefinance)==0:
                # This is the true column name  
                mapping_dict = {column_name_list_in_mapping[i]: entity_list[i] for i in range(len(entity_list))}
                mapped_entity = [mapping_dict[property] if property in mapping_dict else "0" for property in filter_header_row]
                return max_match_row,mapped_entity

            # after apply month_mask, the column_name still doesn't match with that in entity_mapping	
            elif len(duplicate_check)>0: # there is still duplicate property name
                st.error("Detected duplicated column names—— {} in sheet '{}'. Please fix and re-upload.".format(", ".join(f"'{item}'" for item in duplicate_check),sheet_name))
                st.stop()
            elif len(duplicate_check)==0:  # miss some property names              
                max_match=[x for x in filter_header_row if x!=0]
                header_row=filter_header_row
                rest_column_names=[str(x) for x in PL.iloc[max_match_row,:][month_mask] if pd.notna(x) and str(x).upper().strip() not in column_name_list_in_mapping]

        miss_match_column_names = [item for item in column_name_list_in_mapping  if item not in max_match]
	# total missed entities include: missing from P&L, missing(empty) in entity_mapping["column_name"]
        total_missed_entities=entity_mapping[entity_mapping["Column_Name"].str.upper().str.strip().isin(miss_match_column_names)].index.tolist()+entity_without_propertynamefinance
        miss_column_mapping=entity_mapping.loc[total_missed_entities]
        if len(total_missed_entities)>0:
            if len(total_missed_entities)==1:
                st.error("Can't identify the data column for facility: {} in sheet {}. Please add its column name and re-upload.".format(entity_mapping.loc[total_missed_entities[0],"Property_Name"],sheet_name))
                if len(rest_column_names)>0:			
                    st.error("If this facility has a new column name, please re-map it as indicated below.")
                elif len(rest_column_names)==0:
                    st.stop()
            elif len(total_missed_entities)>1:
                st.error("Can't identify the data columns for facilities: {} in sheet {}. Please add their column names and re-upload.".format( ",".join(entity_mapping.loc[total_missed_entities, "Property_Name"]),sheet_name))
                if len(rest_column_names)>0:			
                    st.error("If these facilities have new column names, please re-map them as indicated below.") 
                elif len(rest_column_names)==0:
                    st.stop()
            with st.form(key="miss_match_column_name"):
                for entity_i in total_missed_entities:
                    st.warning("Column name for facility {}".format(entity_mapping.loc[entity_i,"Property_Name"]))
                    miss_column_mapping.loc[entity_i,"Column_Name"]=st.selectbox("Original facility column name: {}".format(\
			entity_mapping.loc[entity_i,"Column_Name"]),[""]+rest_column_names,key=entity_i+"miss_column")
                submitted = st.form_submit_button("Submit")
           
            if submitted:
                if (miss_column_mapping["Column_Name"] == "").any():
                    st.error("Please complete all the mapping.")
                    st.stop()
            
                for entity_i in miss_column_mapping.index: 
                    entity_mapping.loc[entity_i,"Column_Name"]=miss_column_mapping.loc[entity_i,"Column_Name"]     

                column_name_list_in_mapping=[str(x).upper().strip() for x in entity_mapping.loc[entity_list]["Column_Name"]]
                duplicate_check = [item for item in set(column_name_list_in_mapping) if column_name_list_in_mapping.count(item) > 1]
                #st.write("duplicate_check",duplicate_check)	
                if len(duplicate_check)>0:
                    st.error( "The following column has been mapped to more than one facility in sheet '{}'. Please fix and re-upload:".format(sheet_name))
                    st.error(", ".join(f"'{item}'" for item in duplicate_check))
                    st.stop()
		#update header_row
                raw_header_row=list(map(lambda x: str(x).upper().strip() if pd.notna(x) else x,list(PL.iloc[max_match_row,:])))  
                header_row = [item if item in column_name_list_in_mapping else 0 for item in raw_header_row ]

                if len(month_mask)>0: # filter if there are month mask
			
                    header_row=[item if m else 0 for item, m in zip(header_row, month_mask) ]
                duplicate_check = [item for item in set(header_row) if header_row.count(item) > 1 and item!=0]		    
                if len(duplicate_check)>0:
                    st.error("Detected duplicated column names —— {} in sheet '{}'. Please fix and re-upload.".format(", ".join(f"'{item}'" for item in duplicate_check),sheet_name))
                    st.stop()
                elif len([item for item in header_row if item!=0])==len(column_name_list_in_mapping):  # property name column header is unique and match with entity mapping
                    mapping_dict = {column_name_list_in_mapping[i]: entity_list[i] for i in range(len(entity_list))}
                    mapped_entity = [mapping_dict[property] if property in mapping_dict else "0" for property in header_row]
                    Update_File_Onedrive(mapping_path,entity_mapping_filename,entity_mapping,operator,"CSV",None,entity_mapping_str_col)
                    return max_match_row,mapped_entity
            else:
                st.stop()
    else:
        st.stop()    
# no cache
def Read_Clean_PL_Multiple(entity_list,sheet_type,uploaded_file,account_pool,sheet_name):  
    global tenant_account_col
    #st.write("tenant_account_col",tenant_account_col)
    #st.write("account_mapping",account_mapping)
    #check if sheet names in list are same, otherwise, ask user to select correct sheet name.
    #st.write("sheet_type",sheet_type,"account_pool","account_pool",sheet_name)
    if sheet_type=="Sheet_Name_Finance":  
        sheet_type_name="P&L"
    elif sheet_type=="Sheet_Name_Occupancy":
        sheet_type_name="Occupancy"
    elif sheet_type=="Sheet_Name_Balance_Sheet":
        sheet_type_name="Balance"

    # read data from uploaded file
    excel_file = pd.ExcelFile(uploaded_file)

	 
    PL = pd.read_excel(uploaded_file,sheet_name=sheet_name,header=None)
    #st.write("sheet_name",sheet_name,"PL",PL)
    if PL.shape[0]<=1:  # sheet is empty or only has one column
        return pd.DataFrame()
    else:

        # Start checking process  
        tenant_account_col=Identify_Tenant_Account_Col(PL,sheet_name,sheet_type_name,account_pool["Tenant_Account"],tenant_account_col)
        if len(tenant_account_col) > 1:
            # Start with the first column
            tenant_account_col_values = PL.iloc[:, tenant_account_col[0]].fillna('')

            # Iterate over the rest of the columns and combine them
            for col_idx in tenant_account_col[1:]:
                current_col = PL.iloc[:, col_idx].fillna('')
		    
                # Fill missing values in the combined column with values from the current column
                tenant_account_col_values = tenant_account_col_values.where(tenant_account_col_values != '', current_col)

        elif len(tenant_account_col) == 1:
            tenant_account_col_values=PL.iloc[:, tenant_account_col[0]]
        
        tenant_account_col_values = tenant_account_col_values.apply(lambda x: str(int(x)).strip().upper() if pd.notna(x) and isinstance(x, float) else (str(x).strip().upper() if pd.notna(x) else x))
        entity_header_row_number,new_entity_header=Identify_Column_Name_Header(PL,tenant_account_col_values,entity_list,sheet_name) 
	# some tenant account col are in the right side of header, remove these column from tenant_account_col
        if len(tenant_account_col) > 1:
            # Find the index of the first non-'0' in new_entity_header
            first_non_zero_index = next(i for i, value in enumerate(new_entity_header) if value != "0")
            
            # Filter tenant_account_col to keep only indices less than or equal to the first_non_zero_index
            updated_tenant_account_col = [index for index in tenant_account_col if index < first_non_zero_index]
            
            if len(updated_tenant_account_col)<len(tenant_account_col): 
                tenant_account_col_values = PL.iloc[:, updated_tenant_account_col[0]].fillna('')

                # Iterate over the rest of the columns and combine them
                for col_idx in updated_tenant_account_col[1:]:
                    current_col = PL.iloc[:, col_idx].fillna('')
                    # Fill missing values in the combined column with values from the current column
                    tenant_account_col_values = tenant_account_col_values.where(tenant_account_col_values != '', current_col)

	#set tenant_account_col as index of PL
        PL = PL.set_index(tenant_account_col_values)
	    
	#remove row above property header
        PL=PL.iloc[entity_header_row_number+1:,:]

        # remove column without column name, (value in property header that equal to 0)
        non_zero_columns = [val !="0" for val in new_entity_header]
        PL = PL.loc[:,non_zero_columns]    
        PL.columns= [value for value in new_entity_header if value != "0"]
	    
        #remove rows without tenant account
        nan_index=list(filter(lambda x: pd.isna(x) or str(x).strip()=="" or x!=x or x=="nan",PL.index))
        PL.drop(nan_index, inplace=True)
        #set index as str ,strip
        PL.index=map(lambda x:str(x).upper().strip(),PL.index)
        PL=PL.map(lambda x: 0 if pd.isna(x) or isinstance(x, str) or x!=x or x==" " else x)	    
        # don't removes all nan/0, because some property may have no data and need to keep empty
        # remove rows with all nan/0 value
        PL = PL.loc[~PL.apply(lambda x: all(pd.isna(v) or v == 0 or isinstance(v, str) for v in x), axis=1)]

        #st.write("PL before mapping1",PL)	
        # mapping new tenant accounts
        new_tenant_account_list=list(filter(lambda x: str(x).upper().strip() not in list(account_mapping["Tenant_Account"]),PL.index))
        # remove duplicate new account
        new_tenant_account_list=list(set(new_tenant_account_list))    
        if len(new_tenant_account_list)>0:
            #st.write("new_tenant_account_list",new_tenant_account_list)	
            Manage_Account_Mapping(new_tenant_account_list,sheet_name,sheet_type_name)

	    # Update account pool
            if sheet_type=="Sheet_Name_Finance":
                account_pool=account_mapping.copy()
            elif sheet_type=="Sheet_Name_Occupancy":
                account_pool = account_mapping[(account_mapping["Sabra_Account"] == "NO NEED TO MAP")|\
		                (account_mapping["Category"].isin(["Facility Information","Patient Days","Operating Expenses"]))|\
	                        (account_mapping["Sabra_Account"].isin(['T_NURSING_HOURS', 'T_N_CONTRACT_HOURS', 'T_OTHER_HOURS','T_NURSING_LABOR','T_N_CONTRACT_LABOR','T_OTHER_NN_LABOR'])) |\
	                        (account_mapping["Sabra_Second_Account"].isin(['T_NURSING_HOURS', 'T_N_CONTRACT_HOURS', 'T_OTHER_HOURS','T_NURSING_LABOR','T_N_CONTRACT_LABOR','T_OTHER_NN_LABOR']))]
                st.write("Occ pool",account_pool) 
            elif sheet_type=="Sheet_Name_Balance_Sheet":
                account_pool= account_mapping[(account_mapping["Sabra_Account"] == "NO NEED TO MAP")| (account_mapping["Category"]=="Balance Sheet")]	

        #if there are duplicated accounts in P&L, ask for confirming
        # Step 1: Remove all duplicate rows, keeping only unique records based on all column values
        PL.index.name = "Tenant_Account"
        PL = PL.reset_index(drop=False)
        PL=PL.drop_duplicates()
        PL = PL.set_index('Tenant_Account')  
        
        # Step 2: Identify any remaining duplicated indices after removing duplicate rows
        dup_tenant_account_all = PL.index[PL.index.duplicated()].unique()

        # Step 3: Filter out accounts that do not need to be mapped
        dup_tenant_account = [x for x in dup_tenant_account_all \
             if x.upper() not in list(account_mapping[account_mapping["Sabra_Account"] == "NO NEED TO MAP"]["Tenant_Account"])]

        # Step 4: Show error if any duplicated accounts remain after handling duplicates
        if len(dup_tenant_account) > 0:
            st.warning(f"Duplicated accounts detected in {sheet_type_name} sheet '{sheet_name}'. "
             f"Please rectify them to avoid repeated calculations: **{', '.join(dup_tenant_account)}**.")
       
        # Map PL accounts and Sabra account
	# map sabra account with tenant account, groupby sabra account
        #st.write("sheet_type",sheet_type,"PL",PL,"account_pool",account_pool)
        PL=Map_PL_Sabra(PL,entity_list,sheet_type,account_pool) # index are ('ENTITY',"Sabra_Account")
        PL.rename(columns={"value":reporting_month},inplace=True)
        #PL_with_detail.rename(columns={"values":reporting_month},inplace=True) 
    return PL
	
@st.cache_data
def Get_Previous_Months(reporting_month,full_date_header):

    # Convert the reporting_month string to a datetime object
    latest_date = datetime.strptime(reporting_month, "%Y%m")
    month_list = [reporting_month]
    for i in range(previous_monthes_comparison):
        # Subtract i months to get the previous month
        previous_date = latest_date - timedelta(days=latest_date.day, weeks=i*4)
        # Format the date back to the desired string format and append to the list
        month_list.append(previous_date.strftime("%Y%m"))
    month_select=list(filter(lambda x: x in month_list,full_date_header))	
    return month_select

#no cache    
def Read_Clean_PL_Single(entity_i,sheet_type,uploaded_file,wb,account_pool):  
    global tenant_account_col,date_header,select_months_list
    sheet_name=str(entity_mapping.loc[entity_i,sheet_type])
    property_name= str(entity_mapping.loc[entity_i,"Property_Name"] ) 

    if sheet_type=="Sheet_Name_Finance":  
        sheet_type_name="P&L"
    elif sheet_type=="Sheet_Name_Occupancy":
        sheet_type_name="Occupancy"
    elif sheet_type=="Sheet_Name_Balance_Sheet":
        sheet_type_name="Balance"

    # read data from uploaded file
    PL = pd.read_excel(uploaded_file,sheet_name=sheet_name,header=None)	
    #st.write("read PL",PL)
    if PL.shape[0]<=1:  # sheet is empty or only has one column
        return pd.DataFrame()
    if operator in operators_remove_hidden_rowcol:
        ws=wb[sheet_name]
        visible_rows = [row - 1 for row in range(1, PL.shape[0] + 1) if not ws.row_dimensions[row].hidden]

        # Convert Excel column indices (1-based) to Pandas 0-based indices
        visible_cols = [col for col in range(PL.shape[1]) if not ws.column_dimensions.get(get_column_letter(col + 1), None) or not ws.column_dimensions[get_column_letter(col + 1)].hidden]

        # Ensure indices exist before filtering
        if visible_rows and visible_cols:
            PL = PL.iloc[visible_rows, visible_cols]
            # Reset the column indices to be continuous (0, 1, 2, ...)
            PL.columns = range(len(PL.columns))  # Reindex columns
    #st.write("PL",PL)          
    # Start checking process
    with st.spinner("********Start to check facility—'"+property_name+"' in sheet '"+sheet_name+"'********"):
        tenant_account_col=Identify_Tenant_Account_Col(PL,sheet_name,sheet_type_name,account_pool["Tenant_Account"],tenant_account_col)
        
        #st.write("tenant_account_col0",len(tenant_account_col),tenant_account_col)
        if len(tenant_account_col) > 1:
            # Start with the first column
            tenant_account_col_values = PL.iloc[:, tenant_account_col[0]].fillna('')

            # Iterate over the rest of the columns and combine them
            for col_idx in tenant_account_col[1:]:
                current_col = PL.iloc[:, col_idx].fillna('')
                # Fill missing values in the combined column with values from the current column
                tenant_account_col_values = tenant_account_col_values.where(tenant_account_col_values != '', current_col)
        elif len(tenant_account_col) == 1:
            tenant_account_col_values=PL.iloc[:, tenant_account_col[0]]
        tenant_account_col_values = tenant_account_col_values.apply(lambda x: str(int(x)).strip().upper() if pd.notna(x) and isinstance(x, float) else (str(x).strip().upper() if pd.notna(x) else x))
        #st.write("tenant_account_col_values",tenant_account_col_values)
        date_header=Identify_Month_Row(PL,tenant_account_col_values,tenant_account_col[0],sheet_name,sheet_type,date_header)
        #st.write("date_header",date_header)
        if len(date_header[0])==0:
            return pd.DataFrame()
        if all(x=="0" or x==0 for x in date_header[0]):
            st.error("Fail to identify Month/Year header in {} sheet '{}', please add it and re-upload.".format(sheet_type_name,sheet_name))
            st.stop()  
        
	# some tenant account col are in the right side of month header, remove these column from tenant_account_col
        if len(tenant_account_col) > 1:
            # Find the index of the first non-'0' in new_entity_header
            first_non_zero_index = next(i for i, value in enumerate(date_header[0]) if value != "0" and value != 0)

            # Filter tenant_account_col to keep only indices less than or equal to the first_non_zero_index
            updated_tenant_account_col = [index for index in tenant_account_col if index < first_non_zero_index]

            if len(updated_tenant_account_col)<len(tenant_account_col): 
                tenant_account_col_values = PL.iloc[:, updated_tenant_account_col[0]].fillna('')

                # Iterate over the rest of the columns and combine them
                for col_idx in updated_tenant_account_col[1:]:
                    current_col = PL.iloc[:, col_idx].fillna('') 
                    # Fill missing values in the combined column with values from the current column
                    tenant_account_col_values = tenant_account_col_values.where(tenant_account_col_values != '', current_col)

	#set tenant_account_col_values as index of PL
        PL = PL.set_index(tenant_account_col_values)  
        #remove row above date, to prevent to map these value as new accounts
        PL=PL.iloc[date_header[1]+1:,:]
	#remove rows with nan tenant account
        nan_index=list(filter(lambda x:pd.isna(x) or x=="nan" or x=="" or x==" " or x!=x or x==0 ,PL.index))
        PL.drop(nan_index, inplace=True)
        #set index as str ,strip,upper
        PL.index=map(lambda x:str(x).strip().upper(),PL.index)
        #st.write("process PL",PL)    
        # filter columns with month_select
        selected_month_columns = [val in select_months_list for val in date_header[0]]
        #st.write("selected_month_columns",selected_month_columns)
        PL = PL.loc[:,selected_month_columns]   
        PL.columns= [value for value in date_header[0] if value in select_months_list]        
        select_months_list= list(PL.columns)          
        # remove columns with all nan/0 or a combination of nan and 0
        #PL=PL.loc[:,(PL!= 0).any(axis=0)]
        # remove rows with all nan/0 value or a combination of nan and 0 
        #PL = PL.loc[~PL.apply(lambda x: x.isna().all() or (x.fillna(0) == 0).all(), axis=1)]
        PL = PL.loc[~PL.apply(lambda x: all(pd.isna(v) or v == 0 or isinstance(v, str) for v in x), axis=1)]

	# mapping new tenant accounts
        new_tenant_account_list=list(filter(lambda x: x not in list(account_mapping["Tenant_Account"]),PL.index))
        new_tenant_account_list=list(set(new_tenant_account_list))    
        if len(new_tenant_account_list)>0:
            Manage_Account_Mapping(new_tenant_account_list,sheet_name,sheet_type_name)   
            # Update account pool
            if sheet_type=="Sheet_Name_Finance":
                account_pool=account_mapping.copy()
            elif sheet_type=="Sheet_Name_Occupancy":
                account_pool = account_mapping[(account_mapping["Sabra_Account"] == "NO NEED TO MAP")|(account_mapping["Category"] == "Patient Days")|\
	                        (account_mapping["Category"] == "Facility Information")|\
	                        (account_mapping["Sabra_Account"].isin(['T_NURSING_HOURS', 'T_N_CONTRACT_HOURS', 'T_OTHER_HOURS','T_NURSING_LABOR','T_N_CONTRACT_LABOR','T_OTHER_NN_LABOR'])) |\
	                        (account_mapping["Sabra_Second_Account"].isin(['T_NURSING_HOURS', 'T_N_CONTRACT_HOURS', 'T_OTHER_HOURS','T_NURSING_LABOR','T_N_CONTRACT_LABOR','T_OTHER_NN_LABOR']))]	  
                #st.write("Occ pool",account_pool) 
            elif sheet_type=="Sheet_Name_Balance_Sheet":
                account_pool= account_mapping[(account_mapping["Sabra_Account"] == "NO NEED TO MAP")| (account_mapping["Category"]=="Balance Sheet")]	

        #if there are duplicated accounts in P&L, ask for confirming
        # Step 1: Remove all duplicate rows(both of account and responding value are same) keeping only unique records based on column values
        PL.index.name = "Tenant_Account"
        PL = PL.reset_index(drop=False)
        PL=PL.drop_duplicates(subset=["Tenant_Account", reporting_month])
        PL = PL.set_index('Tenant_Account')    
        
        # Step 2: Identify any remaining duplicated indices after removing duplicate rows
        dup_tenant_account_all = PL.index[PL.index.duplicated()].unique()

        # Step 3: Filter out accounts that do not need to be mapped
        dup_tenant_account = [x for x in dup_tenant_account_all \
             if x.upper() not in list(account_mapping[account_mapping["Sabra_Account"] == "NO NEED TO MAP"]["Tenant_Account"])]

        # Step 4: Show error if any duplicated accounts remain after handling duplicates
        if len(dup_tenant_account) > 0:
            st.warning(f"Duplicated accounts detected in {sheet_type_name} sheet '{sheet_name}'. "
             f"Please rectify them to avoid repeated calculations: **{', '.join(dup_tenant_account)}**.")

        
        # Map PL accounts and Sabra account
        PL=Map_PL_Sabra(PL,entity_i,sheet_type,account_pool) 
        #st.write("after Map_PL_Sabr",PL)
    return PL
       

# no cache
def Upload_And_Process(uploaded_file,wb,file_type):
    global tenant_account_col
    Total_PL=pd.DataFrame()
    #Total_PL_detail=pd.DataFrame()
    total_entity_list=list(entity_mapping.index)
    Occupancy_in_one_sheet=[]
    BS_in_one_sheet=[]
    account_pool_full=account_mapping.copy()
    #st.write("account_pool_full",account_pool_full)
    account_pool_patient_days = account_mapping[(account_mapping["Sabra_Account"] == "NO NEED TO MAP")|\
	                        (account_mapping["Category"].isin(["Patient Days","Facility Information","Operating Expenses"]))|\
	                        (account_mapping["Sabra_Account"].isin(['T_NURSING_HOURS', 'T_N_CONTRACT_HOURS', 'T_OTHER_HOURS','T_NURSING_LABOR','T_N_CONTRACT_LABOR','T_OTHER_NN_LABOR'])) |\
	                        (account_mapping["Sabra_Second_Account"].isin(['T_NURSING_HOURS', 'T_N_CONTRACT_HOURS', 'T_OTHER_HOURS','T_NURSING_LABOR','T_N_CONTRACT_LABOR','T_OTHER_NN_LABOR']))]	  
    account_pool_balance_sheet= account_mapping[(account_mapping["Sabra_Account"] == "NO NEED TO MAP")| (account_mapping["Category"]=="Balance Sheet")]	
    #st.write("account_pool_patient_days",account_pool_patient_days,"account_pool_balance_sheet",account_pool_balance_sheet)
    # ****Finance and BS in one excel****
    if file_type=="Finance":
        tenant_account_col=[10000]
        for entity_i in total_entity_list:   # entity_i is the entity code S number
	    # properties are in seperate sheet 
            if entity_mapping.loc[entity_i,"Finance_in_separate_sheets"]=="Y":
                PL=Read_Clean_PL_Single(entity_i,"Sheet_Name_Finance",uploaded_file,wb,account_pool_full)
                if operator!="Ignite":
                    Total_PL = Total_PL.combine_first(PL) if not Total_PL.empty else PL
                else:
                    Total_PL = Total_PL.add(PL, fill_value=0) if not Total_PL.empty else PL

                #st.write("Total_PL",entity_i,Total_PL)
	
	    
	# check census data****************************************************************************************    
        tenant_account_col=[10000]
        for entity_i in total_entity_list:
            sheet_name_finance = str(entity_mapping.loc[entity_i, "Sheet_Name_Finance"])
            sheet_name_occupancy = str(entity_mapping.loc[entity_i, "Sheet_Name_Occupancy"]).strip()
    
            # Check if sheet_name_occupancy is valid, not "nan", and different from sheet_name_finance
            if (sheet_name_occupancy \
                and sheet_name_occupancy.lower() != "nan"\
                and sheet_name_occupancy != sheet_name_finance\
                and entity_mapping.loc[entity_i, "Occupancy_in_separate_sheets"] == "Y" ):

                PL_occ=Read_Clean_PL_Single(entity_i,"Sheet_Name_Occupancy",uploaded_file,wb,account_pool_patient_days) 
                if not PL_occ.empty:
                    if operator!="Ignite":
                        Total_PL=PL_occ.combine_first(Total_PL)
                    else:
                        Total_PL = Total_PL.add(PL_occ, fill_value=0) if not Total_PL.empty else PL_occ
			
	# check BS data******************************		
        tenant_account_col=[10000]
        for entity_i in total_entity_list: 
            if  entity_mapping.loc[entity_i,"BS_separate_excel"]=="N": 
                sheet_name_finance=str(entity_mapping.loc[entity_i,"Sheet_Name_Finance"])
                sheet_name_balance=str(entity_mapping.loc[entity_i,"Sheet_Name_Balance_Sheet"])
                if not pd.isna(sheet_name_balance) \
                       and sheet_name_balance!=" " \
                       and sheet_name_balance!="nan" \
                       and sheet_name_balance!=sheet_name_finance \
                       and entity_mapping.loc[entity_i,"Balance_in_separate_sheets"]=="Y":
                    PL_BS=Read_Clean_PL_Single(entity_i,"Sheet_Name_Balance_Sheet",uploaded_file,wb,account_pool_balance_sheet)
                    if PL_BS.shape[0]>0:
                        Total_PL=PL_BS.combine_first(Total_PL)
        
	# All the properties are in one sheet	
        sheet_list_finance_in_onesheet = entity_mapping[entity_mapping["Finance_in_separate_sheets"]=="N"]["Sheet_Name_Finance"].unique()
        if len(sheet_list_finance_in_onesheet)>0:
            #st.write("sheet_list_finance_in_onesheet",sheet_list_finance_in_onesheet)
            for sheet_name_finance_in_onesheet in sheet_list_finance_in_onesheet:
                tenant_account_col=[10000]
                entity_list_finance_in_onesheet=entity_mapping.index[entity_mapping["Sheet_Name_Finance"]==sheet_name_finance_in_onesheet].tolist()
                PL=Read_Clean_PL_Multiple(entity_list_finance_in_onesheet,"Sheet_Name_Finance",uploaded_file,account_pool_full,sheet_name_finance_in_onesheet)
                if operator!="Ignite":               
                    Total_PL = Total_PL.combine_first(PL) if not Total_PL.empty else PL
                else:
                    Total_PL = Total_PL.add(PL, fill_value=0) if not Total_PL.empty else PL


	# census
        sheet_list_occupancy_in_onesheet = entity_mapping[(entity_mapping["Occupancy_in_separate_sheets"]=="N")&(~pd.isna(entity_mapping["Sheet_Name_Occupancy"]))&(entity_mapping["Sheet_Name_Occupancy"]!="nan")]["Sheet_Name_Occupancy"].unique()
        if len(sheet_list_occupancy_in_onesheet)>0:
            for sheet_name_occupancy_in_onesheet in sheet_list_occupancy_in_onesheet:
                tenant_account_col=[10000]
                entity_list_occupancy_in_onesheet=entity_mapping.index[entity_mapping["Sheet_Name_Occupancy"]==sheet_name_occupancy_in_onesheet].tolist()	
                PL_Occ=Read_Clean_PL_Multiple(entity_list_occupancy_in_onesheet,"Sheet_Name_Occupancy",uploaded_file,account_pool_patient_days,sheet_name_occupancy_in_onesheet)
                if PL_Occ.shape[0]>0:
                    if operator!="Ignite": 
                        Total_PL=PL_Occ.combine_first(Total_PL)
                    else:
                        Total_PL = Total_PL.add(PL_Occ, fill_value=0) if not Total_PL.empty else PL_Occ
		    
	# balance sheet
        sheet_list_bs_in_onesheet = entity_mapping[(entity_mapping["Balance_in_separate_sheets"]=="N")&(entity_mapping["BS_separate_excel"]=="N")&(~pd.isna(entity_mapping["Sheet_Name_Balance_Sheet"]))&(entity_mapping["Sheet_Name_Balance_Sheet"]!="nan")]["Sheet_Name_Balance_Sheet"].unique()
        if len(sheet_list_bs_in_onesheet)>0:
            for sheet_name_bs_in_onesheet in sheet_list_bs_in_onesheet:
                tenant_account_col=[10000]
                entity_list_bs_in_onesheet=entity_mapping.index[entity_mapping["Sheet_Name_Balance_Sheet"]==sheet_name_bs_in_onesheet].tolist()	
                PL_BS=Read_Clean_PL_Multiple(entity_list_bs_in_onesheet,"Sheet_Name_Balance_Sheet",uploaded_file,account_pool_balance_sheet,sheet_name_bs_in_onesheet)
                if PL_BS.shape[0]>0:
                    Total_PL=PL_BS.combine_first(Total_PL)
		    
    elif file_type=="BS":
        tenant_account_col=[10000]
        for entity_i in total_entity_list: 
            if entity_mapping.loc[entity_i,"Balance_in_separate_sheets"]=="Y":
                PL_BS=Read_Clean_PL_Single(entity_i,"Sheet_Name_Balance_Sheet",uploaded_file,wb,account_pool_balance_sheet)             
                Total_PL = PL_BS.combine_first(Total_PL) if not Total_PL.empty else PL_BS

        sheet_list_bs_in_onesheet = entity_mapping[(entity_mapping["Balance_in_separate_sheets"]=="N")&(~pd.isna(entity_mapping["Sheet_Name_Balance_Sheet"]))&(entity_mapping["Sheet_Name_Balance_Sheet"]!="nan")]["Sheet_Name_Balance_Sheet"].unique()
        if len(sheet_list_bs_in_onesheet)>0:
            for sheet_name_bs_in_onesheet in sheet_list_bs_in_onesheet:
                tenant_account_col=[10000]
                entity_list_bs_in_onesheet=entity_mapping.index[entity_mapping["Sheet_Name_Balance_Sheet"]==sheet_name_bs_in_onesheet].tolist()	
                PL_BS=Read_Clean_PL_Multiple(entity_list_bs_in_onesheet,"Sheet_Name_Balance_Sheet",uploaded_file,account_pool_balance_sheet,sheet_name_bs_in_onesheet)  
                Total_PL = PL_BS.combine_first(Total_PL) if not Total_PL.empty else PL_BS

    Total_PL = Total_PL.sort_index()  #'ENTITY',"Sabra_Account" are the multi-index of Total_Pl
    #st.write("Total_PL",Total_PL)
    return Total_PL
def Download_PL_Sample():
    PL_sample_filename = "{}_P&L_sample.xlsx".format(operator)
    
    # Fetch data from OneDrive
    #PL_sample = Read_File_From_Onedrive(mapping_path, PL_sample_filename, "XLSX")
    
    #if PL_sample is not False:
        # Create a BytesIO buffer to hold the Excel data
        #output = BytesIO()
        #with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        #    PL_sample.to_excel(writer, index=False)
        #download_file = output.getvalue()

        # Return the download button with the Excel file data
        #st.download_button(label="Download P&L sample",data=download_file,file_name=PL_sample_filename,mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" )
    #else:
        #st.write("P&L sample is not found. Please contact sli@sabrahealth.com to get it.")
#----------------------------------website widges------------------------------------
config = Read_File_From_Onedrive(mapping_path, "config.yaml","YAML")
# Creating the authenticator object
if config:
    authenticator = Authenticate(
        config['credentials'],
        config['cookie']['name'], 
        config['cookie']['key'], 
        config['cookie']['expiry_days'],
        config['preauthorized'])
else:
    st.write("Failed to read configuration from OneDrive.")

# set button status
button_initial_state={"forgot_password_button":False,"forgot_username_button":False,"submit_report":False}

if 'clicked' not in st.session_state:
    st.session_state.clicked = button_initial_state
	

# login widget
col1,col2=st.columns(2)
with col1:
    authenticator.login('Login',config,'main')
if st.session_state["authentication_status"] is False:
    st.error('Username/password is incorrect')
#------------------------------------------operator account----------------------------------------------------------
elif st.session_state["authentication_status"] and st.session_state["operator"]!="Sabra":
    operator_email = st.session_state['email']  # Accessing the email
    
    operator=st.session_state["operator"]
    st.title(operator)
    menu=["Upload P&L","Manage Mapping","Instructions","Edit Account","Logout"]
    if "operator_choice" not in st.session_state:
        st.session_state["operator_choice"] = menu[0]  # Set default value
    choice=st.sidebar.selectbox("Menu", menu,key="operator_choice",index=menu.index(st.session_state["operator_choice"]))
    
    if choice=="Upload P&L":
        if current_month<10:
            current_date=str(current_year)+"0"+str(current_month)
        else:
            current_date=str(current_year)+str(current_month)
      
        BPC_pull,entity_mapping,account_mapping=Initial_Mapping(operator)
        if not BPC_pull.empty:
            try:
                reporting_month = BPC_pull["Reporting_Month"].dropna().iloc[0] if not BPC_pull["Reporting_Month"].dropna().empty else None
            except: 
                reporting_month = None
        else:
            reporting_month = None		
	#st.write("reporting_month",reporting_month)
        months_range = list(month_map.keys())
        if 'selected_year' not in st.session_state:
            if reporting_month is not None and reporting_month[0:2]=='20':
                st.session_state.selected_year = int(reporting_month[0:4])
            else:
                st.session_state.selected_year = current_year    
        if "email_sent" not in st.session_state:
            st.session_state.email_sent = False
        #st.write("current_month",current_month)  	    
	
        reporting_month_label=True  
        tenant_account_col=[10000]
        date_header=[[0],0,[]]
        select_months_list=[]
        # Calculate the list of years and their indices
        years_range = list(range(current_year, current_year - 2, -1))
        # Calculate the list of months and their indices
        months_range = list(month_map.keys())
        if "Y" in entity_mapping["BS_separate_excel"][(pd.notna(entity_mapping["BS_separate_excel"]))&(entity_mapping["DATE_SOLD_PAYOFF"]=="N")].values:             
            BS_separate_excel="Y"
        else:
            BS_separate_excel="N"
        if BS_separate_excel=="N":
            with st.form("upload_form", clear_on_submit=True):
                st.subheader("Select reporting month:") 
                col1,col2=st.columns([1,1])
                with col1:
                    selected_year = st.selectbox("Year", years_range,index=years_range.index(st.session_state.selected_year))
                    st.session_state.selected_year = selected_year
                   
                with col2:    
                    selected_month = st.selectbox("Month", months_range)
                    
                with col1:
                    st.write("Upload P&L:")
                    uploaded_finance=st.file_uploader("Upload your finance file",label_visibility="hidden",type={"xlsx"},accept_multiple_files=False,key="Finance_upload")
                with col2:
                    st.write("Other Documents:")
                    uploaded_other_docs = st.file_uploader( "Optional", 
                                         type=["csv", "pdf", "xlsm", "xlsx", "xls", "zip"], 
                                         accept_multiple_files=True, 
                                         key="Other docs")

                submitted = st.form_submit_button("Calculate & Review (Not Yet Submitted)")
                
                if submitted:
	            # clear cache for every upload
                    st.cache_data.clear()
                    st.cache_resource.clear()
                    st.session_state.clicked = button_initial_state
                    st.session_state.selected_year = selected_year
                    st.session_state.email_sent = False
                    st.session_state.email_body_for_Sabra = ""
        elif BS_separate_excel=="Y":	 
            with st.form("upload_form", clear_on_submit=True):
                st.subheader("Select reporting month:") 
                col1,col2,col3=st.columns([1, 1, 1])
                with col1:
                    selected_year = st.selectbox("Year", years_range,index=years_range.index(st.session_state.selected_year))
                with col2:    
                    selected_month = st.selectbox("Month", months_range)  
                col1, col2, col3 = st.columns([1, 1, 1])
                with col1:
                    st.write("Upload P&L:")
                    uploaded_finance=st.file_uploader("upload financial",type={"xlsx"},accept_multiple_files=False,label_visibility="hidden",key="Finance_upload")
                with col2:
                    st.write("Upload Balance Sheet:")
                    uploaded_BS=st.file_uploader("uploadBS",type={"xlsx"},accept_multiple_files=False,key="BS_upload",label_visibility="hidden")
                with col3:
                    st.write("Other Documents:")
                    uploaded_other_docs=st.file_uploader("Optional",type=["csv","pdf","xlsm","xlsx","xls"],accept_multiple_files=True,key="Other docs")
                submitted = st.form_submit_button("Calculate & Review (Not Yet Submitted)")
                if submitted:
	            # clear cache for every upload
                    st.cache_data.clear()
                    st.cache_resource.clear()
                    st.session_state.clicked = button_initial_state
                    st.session_state.selected_year = selected_year
                    st.session_state.email_body_for_Sabra = ""
        reporting_month_display=str(selected_month)+" "+str(selected_year)
        reporting_month=str(selected_year)+month_map[selected_month]
        if operator=="Ensign_Aspen":
            SHAREPOINT_FOLDER = "Asset Management/01_Operators/Ensign/Financials & Covenant Analysis/_Facility Financials/{}/.{} {}".format(str(selected_year), month_map[selected_month], selected_month)  
        else:
            SHAREPOINT_FOLDER = "Asset Management/01_Operators/{}/Financials & Covenant Analysis/_Facility Financials/{}/.{} {}".format(operator, str(selected_year), month_map[selected_month], selected_month)  
      
        if reporting_month>=current_date:
            st.error("The reporting month should precede the current month.")
            st.stop()	
        entity_mapping=entity_mapping.loc[((entity_mapping["DATE_ACQUIRED"]<=reporting_month) & ((entity_mapping["DATE_SOLD_PAYOFF"]=="N")|(entity_mapping["DATE_SOLD_PAYOFF"]>=reporting_month))),]
        if entity_mapping.empty:
            st.error("The reporting month is not valid as it either exceeds the sold date or precedes the acquisition date for the properties.")
            st.stop()
        #uploading all the files to sharepoint
        upload_list=[]
        upload_filename_list=[]
        if uploaded_finance:
            upload_list.append(uploaded_finance)
            upload_filename_list.append("{}-{}_{}_P&L_uploaded_{}-{}.xlsx".format(reporting_month[4:6],reporting_month[0:4],operator,current_month,current_day))
        if BS_separate_excel=="Y" and uploaded_BS:
            upload_list.append(uploaded_BS)
            upload_filename_list.append("{}-{}_{}_BS_uploaded_{}-{}.xlsx".format(reporting_month[4:6],reporting_month[0:4],operator,current_month,current_day))
        if uploaded_other_docs:
            for file in uploaded_other_docs:
                upload_list.append(file)
                upload_filename_list.append("{}-{}_{}".format(reporting_month[4:6],reporting_month[0:4],file.name))	    
        if upload_filename_list:
            success,upload_message  = Upload_To_Sharepoint(upload_list, SHAREPOINT_FOLDER,upload_filename_list)
            if success:
                st.success("{} files for {} uploaded successfully.".format(len(upload_list),reporting_month_display))
            elif not success and upload_message!=[]:
                st.session_state.email_body_for_Sabra+=f"""
	        <p><strong>{len(upload_message)}</strong> files failed to upload:</p>  
                <ul>  
                    {''.join(f'<li>{file}</li>' for file in upload_message)}  
                </ul>
	        """
            elif upload_message==[]:
                st.session_state.email_body_for_Sabra+=f"""<p><strong>Error during SharePoint upload process. Files are not uploaded</p> """

            success_onedrive,upload_message_onedrive=Upload_To_Onedrive(upload_list,"{}/{}".format(PL_path,operator),upload_filename_list)
            if not success_onedrive and upload_message!=[]:
                 st.session_state.email_body_for_Sabra+=f"""
	        <p><strong>{len(upload_message_onedrive)}</strong> files failed to upload to onedrive:</p>  
                <ul>  
                    {''.join(f'<li>{file}</li>' for file in upload_message_onedrive)}  
                </ul>
	        """
 
        col1, col2 = st.columns([1,3])  
        if 'uploaded_finance' in locals() and uploaded_finance:
            with col1:
                st.markdown("✔️ :green[P&L selected]")
        else:   
            if uploaded_other_docs: 
                st.error("You have only uploaded ancillary files without any monthly reporting data.")
                unique_asset_managers = entity_mapping['Asset_Manager'].unique()
                receiver = operator_email.split(",") + ["twarner@sabrahealth.com", "sli@sabrahealth.com"]  
                receiver.extend(unique_asset_managers)	 
                if not st.session_state.email_sent:
                    st.session_state.email_sent = True
            else:
                st.markdown(":red[P&L is not uploaded ]")
            st.stop()
        
        if "Y" in entity_mapping["BS_separate_excel"][pd.notna(entity_mapping["BS_separate_excel"])].values:                     
            BS_separate_excel="Y"
            if 'uploaded_BS' in locals() and uploaded_BS:
                with col2:
                    st.markdown("✔️ :green[Balance sheet selected]")
            else:
                with col2:
                    st.markdown("❌ :red[Balance sheet is not uploaded ]")
                if uploaded_other_docs: 
                    st.warning("You have only uploaded ancillary files without any monthly reporting data. Please upload Balance Sheet if you want to upload monthly reporting data.")
                st.stop()
        else:
            BS_separate_excel="N"
	
	# select_months_list contain the monthes that need to be compared for history data,if it is [], means no need to compare
        if any(entity_mapping["Finance_in_separate_sheets"]=="N") or previous_monthes_comparison==0:
            select_months_list=[reporting_month]
        else:
            select_months_list =sorted([x for x in BPC_pull.columns if x <reporting_month],reverse=True)
            if len(select_months_list)>=previous_monthes_comparison:
                select_months_list=select_months_list[:previous_monthes_comparison]+[reporting_month]  
            else:
                select_months_list.append(reporting_month) 
        with st.spinner("Processing report... Please wait and click 'Confirm' when it's ready. Do not leave this page until you complete the process."):
            if BS_separate_excel=="N":  # Finance/BS are in one excel
                entity_mapping,finance_wb=Check_Sheet_Name_List(uploaded_finance,"Finance")	 
                Total_PL=Upload_And_Process(uploaded_finance,finance_wb,"Finance")

            elif BS_separate_excel=="Y": # Finance/BS are in different excel 
                entity_mapping,finance_wb=Check_Sheet_Name_List(uploaded_finance,"Finance")
                entity_mapping,bs_wb=Check_Sheet_Name_List(uploaded_BS,"BS")

                # process Finance 
                Total_PL=Upload_And_Process(uploaded_finance,finance_wb,"Finance")

	        # process BS 
                Total_BS=Upload_And_Process(uploaded_BS,bs_wb,"BS")
	        # combine Finance and BS
                Total_PL=Total_BS.combine_first(Total_PL)

            if len(Total_PL.columns)==1:
                Total_PL.columns=[reporting_month]

            elif len(Total_PL.columns)>1 and BPC_pull.shape[0]>0:  # there are previous months in P&L
                diff_BPC_PL=Compare_PL_Sabra(Total_PL,reporting_month)


	
	    # 1 Summary
            total_email_body=View_Summary()
            # Define the button and handle the click event
            if st.button(f'Confirm and upload {reporting_month_display} reporting', key='reporting_month', help="Click and wait a few seconds for the confirmation message."):
                st.session_state.clicked['submit_report'] = True

            # Perform the upload action here and check for discrepancies
            if st.session_state.clicked['submit_report']:
                Submit_Upload(total_email_body,SHAREPOINT_FOLDER)
                # Discrepancy of Historic Data
                #if len(Total_PL.columns) > 1 and BPC_pull.shape[0] > 0:
                    #with st.expander("Discrepancy for Historic Data", expanded=True):
                        #ChangeWidgetFontSize('Discrepancy for Historic Data', '25px')
                        #View_Discrepancy()

    elif choice=="Manage Mapping":
        BPC_pull,entity_mapping,account_mapping=Initial_Mapping(operator)
        with st.expander("Manage Property Mapping" ,expanded=True):
            ChangeWidgetFontSize('Manage Property Mapping', '25px')
            entity_mapping=Manage_Entity_Mapping(operator)
        with st.expander("Manage Account Mapping",expanded=True):
            ChangeWidgetFontSize('Manage Account Mapping', '25px')
            col1,col2=st.columns(2)
            with col1:
                new_tenant_account=st.text_input("Enter new account and press Enter to apply. Use commas to separate them if there are multiple accounts.")
                if new_tenant_account:
                    new_tenant_account_list=list(set(map(lambda x:x.strip(),new_tenant_account.split(",") )))
                    duplicate_accounts=list(filter(lambda x:x.upper() in list(account_mapping['Tenant_Account']),new_tenant_account_list))
                   
                    if len(duplicate_accounts)>1:
                        st.write("{} are already existed in mapping list and will be skip.".format(",".join(duplicate_accounts)))
                    elif len(duplicate_accounts)==1:
                        st.write("{} is already existed in mapping list and will be skip.".format(duplicate_accounts[0]))
		
		    # remove duplicated accounts
                    new_tenant_account_list=list(set(new_tenant_account_list) - set(duplicate_accounts))
                    if len(new_tenant_account_list)==0:
                        st.stop()
                    Manage_Account_Mapping(new_tenant_account_list)
                    #Update_File_Onedrive(mapping_path,account_mapping_filename,account_mapping,operator,"XLSX",None,account_mapping_str_col)

    elif choice=='Instructions':
        # insert Video
        video=Read_File_From_Onedrive(mapping_path,"Sabra App video.mp4","VIDEO")
        st.video(video, format="video/mp4", start_time=0)
	    
    elif choice=="Edit Account": 
	# update user details widget
        try:
            authenticator.update_user_details(st.session_state["username"], 'Update user details',config)

        except Exception as e:
            st.error(e)

    elif choice=="Logout":
        authenticator.logout('Logout', 'main')
# ---------------------------------------- Sabra Admin account----------------------------------------------------------	    
elif st.session_state["authentication_status"] and st.session_state["operator"]=="Sabra":
    operator_list=Read_File_From_Onedrive(mapping_path,operator_list_filename,"CSV")

    menu=["Review Monthly reporting","Review New Mapping","Edit Account","Register","Logout"]
    if "Sabra_choice" not in st.session_state:
        st.session_state["Sabra_choice"] = menu[0]  # Set default value
    choice=st.sidebar.selectbox("Menu", menu,key="Sabra_choice",index=menu.index(st.session_state["Sabra_choice"]))



    if choice=="Edit Account":
	# update user details widget
        try:
            if authenticator.update_user_details(st.session_state["username"], 'Update user details',config):
                st.success('Updated successfully')
        except Exception as e:
            st.error(e)
    
    elif choice=="Register":
        #st.write("operator_list",operator_list)
        col1,col2=st.columns(2)
        with col1:
            operator= st.selectbox('Select Operator',(operator_list["Operator"]))
        try:
            if authenticator.register_user('Register for '+operator, operator, config, preauthorization=False):
                st.success('Registered successfully')
        except Exception as e:
            st.error(e)
		
    elif choice=="Logout":
        authenticator.logout('Logout', 'main')

    elif choice=="Review Monthly reporting":
        def Create_EPM_Formula(summary,upload_data,selected_indices):
            def colnum_letter(col_number):
                letter = ""
                col_number+=1
                while col_number > 0:
                    col_number, remainder = divmod(col_number - 1, 26)
                    letter = chr(65 + remainder) + letter
                return letter 

            entity_mapping = (
                Read_File_From_Onedrive(mapping_path, entity_mapping_filename, "CSV", entity_mapping_str_col)
                .reset_index(drop=True))

            # Filter where "Operator" matches any value in the operator list
            operator_list=upload_data["Operator"].unique()
            entity_mapping = entity_mapping[entity_mapping["Operator"].isin(operator_list)]
 
            upload_data=upload_data.fillna(0)
            upload_data['Amount']=upload_data['Amount'].astype(float)
        
            upload_data=upload_data.merge(entity_mapping[["ENTITY","Operator","GEOGRAPHY","LEASE_NAME","FACILITY_TYPE","INV_TYPE"]],on=["ENTITY","Operator"],how="left")
            
            entity_list=upload_data["ENTITY"].unique()
            upload_data["EPM_Formula"]=""
            upload_data["Upload_Check"]=""
            row_size=upload_data.shape[0]
            col_name_list=list(upload_data.columns)
            time_col_letter=colnum_letter(col_name_list.index("TIME"))
            entity_col_letter=colnum_letter(col_name_list.index("ENTITY"))
            account_col_letter=colnum_letter(col_name_list.index("Sabra_Account"))
            facility_col_letter=colnum_letter(col_name_list.index("FACILITY_TYPE"))
            state_col_letter=colnum_letter(col_name_list.index("GEOGRAPHY"))
            leasename_col_letter=colnum_letter(col_name_list.index("LEASE_NAME"))
            inv_col_letter=colnum_letter(col_name_list.index("INV_TYPE"))
            data_col_letter=colnum_letter(col_name_list.index("Amount"))    
            EPM_Formula_col_letter=colnum_letter(col_name_list.index("EPM_Formula"))
            upload_Check_col_letter=colnum_letter(col_name_list.index("Upload_Check"))
            
            for r in range(2,row_size+2):
                upload_formula="""=@EPMSaveData({}{},"finance",{}{},{}{},{}{},{}{},{}{},{}{},{}{},"D_INPUT","F_NONE","USD","PERIODIC","ACTUAL")""".\
		    format(data_col_letter,r,time_col_letter,r,entity_col_letter,r,account_col_letter,r,facility_col_letter,r,state_col_letter,r,leasename_col_letter,r,inv_col_letter,r)
                upload_data.loc[r-2,"EPM_Formula"]=upload_formula
                upload_check_formula="={}{}={}{}".format(EPM_Formula_col_letter,r,data_col_letter,r)
                upload_data.loc[r-2,"Upload_Check"]=upload_check_formula
            upload_data["""="Consistence check:"&AND({}2:{}{})""".format(upload_Check_col_letter,upload_Check_col_letter,row_size+1)]=""
            return upload_data         


	    
        st.subheader("Uploading Summary")
        data=Read_File_From_Onedrive(master_template_path,monthly_reporting_filename,"CSV")

        if data is False or data.empty:
            st.warning("The master template is empty or invalid. Please check the file in onedrive.")
        else:
            data=data[list(filter(lambda x:"Unnamed" not in x and 'index' not in x ,data.columns))]
            data["Upload_Check"]=""

            data["TIME"]=data["TIME"].apply(lambda x: "{}.{}".format(str(x)[0:4],month_abbr[int(str(x)[4:6])]))

            # show uploading summary
            summary=data[["TIME","Operator","Latest_Upload_Time"]].drop_duplicates()
            summary['Latest_Upload_Time'] = summary['Latest_Upload_Time'].apply(
    lambda x: pd.to_datetime(x, format="%m/%d/%Y %H:%M", errors='coerce') 
    if len(x.split(' ')[0].split('/')) == 3 else pd.to_datetime(x, format="%Y-%m-%d %H:%M", errors='coerce'))
            summary = summary.sort_values(by='Latest_Upload_Time', ascending=False, na_position='last')
            
            summary = summary.reset_index(drop=True)  # Reset index to create a numeric index
            summary["Index"] = summary.index + 1      # Add a column with 1-based indices
            with st.container():
                st.dataframe(
                        summary[["Index","Operator","TIME","Latest_Upload_Time"]],
                        column_config={
			"Index":"Index No.",
                        "TIME": "Reporting month",
                        "Latest_Upload_Time": "Latest submit time"},
                        hide_index=True)

            
	    # download reporting data with EPM formula
            st.write("")
            st.subheader("Download selected reports with EPM Formula")   
            col1,col2=st.columns([1,2])
            with col1:
                selected_indices = st.multiselect(
                    "Select index(es) to download reports",
                    options=summary["Index"].tolist())

            # Use session state to manage button clicks
            if "show_download" not in st.session_state:
                st.session_state.show_download = False

            # Button to select index
            if st.button("Confirm selections"):
                if selected_indices:
                    st.session_state.show_download = True
                else:
                    st.warning("Please select at least one index No. to download.")

                # Display the download button if index are selected
                if st.session_state.show_download:
                    selected_reports = summary[summary["Index"].isin(selected_indices)].drop(columns=["Latest_Upload_Time"])
                    filtered_data = data.merge(selected_reports, on=["TIME", "Operator"])
                    upload_data_EPM_fomula=Create_EPM_Formula(summary,filtered_data,selected_indices)
                    # Convert result_data to CSV
                    csv = upload_data_EPM_fomula.to_csv(index=False).encode('utf-8')
                    st.download_button(
                        label="Download reporting data",
                        data=csv,
                        file_name="Operator_reporting_data.csv",
                        mime="text/csv")
